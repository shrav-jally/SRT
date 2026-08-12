"""
Excel Writer Module

Writes mapped financial data to the Excel template file.
Reads the existing template structure and fills in the values
in the "values" column (column C) for all sheets.

After writing extracted values, applies Excel formulas for
calculated/highlighted fields (totals, subtotals, EBITDA, etc.)
so they auto-compute from the extracted component values.

Template layout:
  Balance Sheet: Col A (category), Col B (label), Col C (values)
  P&L:          Col A (category), Col B (label), Col C (values)
  Cash Flow:    Col A (category), Col B (label), Col C (values)
"""

import logging
import os
import re
from copy import copy
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .data_mapper import MappingResult
from .text_table_extractor import ExtractedTable

logger = logging.getLogger(__name__)


# ============================================================================
# CONSTANTS
# ============================================================================

VALUES_COLUMN = 3  # Column C — the "values" column in all sheets


# ============================================================================
# DEFAULT ZERO ROWS — REMOVED
# ============================================================================
# Previously, these sets defined row numbers that would default to 0 if no
# value was extracted. This has been REMOVED per policy:
#   "Only write 0 if the PDF explicitly says 0/Nil/blank/— for that item.
#    If the item is simply NOT in the PDF, leave the Excel cell blank."
#
# The old sets are kept as empty sets for backward compatibility but will
# produce no default-zero writes.

BS_DEFAULT_ZERO_ROWS = set()        # Removed — no auto-zero for absent items
PL_DEFAULT_ZERO_ROWS_FULL = set()   # Removed — no auto-zero for absent items
CF_DEFAULT_ZERO_ROWS = set()        # Removed — no auto-zero for absent items


# ============================================================================
# BALANCE SHEET: Section row ranges for disambiguation
# ============================================================================

# Some labels appear in multiple sections (e.g., "(i) Borrowings" in both
# Non-current and Current liabilities). These ranges let us search within
# the correct section based on the mapping result's section field.

BS_SECTION_ROW_RANGES = {
    "Non-current assets": (3, 16),
    "Current assets": (20, 29),
    "Non-current liabilities": (42, 50),
    "Current liabilities": (53, 59),
}

# Top-level/total items that don't belong to a nested section.
# Their row numbers are fixed in the template.
BS_KNOWN_ITEM_ROWS = {
    "Total Non Current Assets": 17,
    "Total Fixed assets": 18,
    "Total Current Assets": 30,
    "Total Assets": 31,
    "Equity Share Capital": 34,
    "Other Equity": 35,
    "Total Equity": 36,
    "Profit for the year": 37,
    "Change in FCTR": 38,
    "NCI share of loss": 39,
    "Total Non-current Liabilities": 51,
    "Total Current Liabilities": 60,
    "Total Liabilities": 61,
    "Total Equity and Liabilities": 62,
    "Total debt": 63,
}


# ============================================================================
# P&L: Section row ranges
# ============================================================================

PL_SECTION_ROW_RANGES = {
    "Income": (3, 5),
    "Taxes": (8, 10),
    "Profit/Loss": (13, 17),
    "IV. Expenses": (22, 32),
}

PL_KNOWN_ITEM_ROWS = {
    "EBITDA": 19,
    "EBIT": 20,
}


# ============================================================================
# Cash Flow: Section row ranges
# ============================================================================

CF_SECTION_ROW_RANGES = {
    "Cash Flow": (3, 5),
    "Other Financial Information": (7, 25),
}


# ============================================================================
# EXCEL FORMULAS FOR CALCULATED/HIGHLIGHTED FIELDS
# ============================================================================

# These formulas are written to highlighted rows after all extracted values
# are in place. They reference the cells in column C where component values
# were written, so totals auto-compute from their components.

BS_FORMULAS = {
    # --- Parent row formulas (sum of children) ---
    # These prevent double-counting: the parent row shows the sub-total,
    # and top-level totals reference the parent instead of summing children.
    10: "=C11+C12+C13+C14",   # (h) Financial Assets (NC) = Investments + Trade receivables + Loans + Others
    21: "=SUM(C22:C27)",      # (b) Financial Assets (Current) = Investments through Others
    42: "=C43+C44+C47",       # (a) Financial Liabilities (NC) = Borrowings + Trade Payables + Other fin liab
    44: "=C45+C46",           # (ii) Trade Payables (NC) = micro/small + other creditors
    53: "=C54+C55+C56",       # (a) Financial Liabilities (Current) = Borrowings + Trade payables + Other fin liab

    # --- Top-level totals (reference parent rows to avoid double-counting) ---
    17: "=C3+C4+C5+C6+C7+C8+C9+C10+C15+C16",  # Total NC Assets (leaf items + parent row 10)
    18: "=C17",                                   # Total Fixed assets (= Total NC Assets)
    30: "=C20+C21+C28+C29",                      # Total Current Assets (leaf items + parent row 21)
    31: "=C17+C30",                               # Total Assets = NC Assets + Current Assets
    36: "=C34+C35",                               # Total Equity = Share Capital + Other Equity
    51: "=C42+C48+C49+C50",                      # Total NC Liabilities (parent row 42 + leaf items)
    60: "=C53+C57+C58+C59",                      # Total Current Liabilities (parent row 53 + leaf items)
    61: "=C51+C60",                               # Total Liabilities = NC Liab + Current Liab
    62: "=C36+C61",                               # Total Equity and Liabilities
    63: "=C43+C54",                               # Total debt = NC Borrowings + Current Borrowings
}

PL_FORMULAS = {
    5:  "=C3+C4",             # III. Total Income (I + II)
    10: "=C8+C9",             # VIII. Tax expense = Current + Deferred
    16: "=C17+C15",           # XIII. Profit after taxes (IX + XII)
    17: "=C32-C10",           # IX. Profit from continuing operations (VII-VIII)
    19: "=C32+C27+C26",      # EBITDA = PBT + D&A + Finance costs
    20: "=C32+C26",           # EBIT = PBT + Finance costs
    29: "=SUM(C22:C28)",     # Total expenses
    30: "=C5-C29",            # V. Profit before exceptional items and tax (III-IV)
    32: "=C30-C31",           # VII. Profit before tax (V-VI)
}

# PL_DEFAULT_ZERO_ROWS removed — per policy, only write 0 if PDF explicitly says 0/Nil/blank


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def _normalize_for_match(s: str) -> str:
    """Normalize a string for comparison."""
    s = s.lower().strip()
    s = re.sub(r'[\(\)\[\],\.\-:/]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _find_row_in_range(
    ws,
    label: str,
    start_row: int,
    end_row: int,
    search_col: int = 2,  # Column B by default
) -> Optional[int]:
    """
    Find the row number for a label within a specific row range.

    Uses normalized matching (lowercase, punctuation removed) for robustness.

    Returns:
        Row number or None if not found.
    """
    label_norm = _normalize_for_match(label)

    for row in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row, column=search_col).value
        if cell_value is None:
            continue
        cell_norm = _normalize_for_match(str(cell_value))
        if cell_norm == label_norm:
            return row
        # Partial match (one contains the other)
        if label_norm and cell_norm and len(label_norm) > 5:
            if label_norm in cell_norm or cell_norm in label_norm:
                return row

    return None


def _find_row_for_item(
    ws,
    template_item: str,
    section: str,
    section_ranges: dict,
    known_items: dict,
) -> Optional[int]:
    """
    Find the row for a template item, using section-based row ranges
    for disambiguation when the same label appears in multiple sections.

    Strategy:
    1. If item is in known_items dict, use the hardcoded row
    2. If section is in section_ranges, search within that range
    3. Fall back to searching all of column B

    Returns:
        Row number or None if not found.
    """
    # Check known items first (totals, top-level items)
    if template_item in known_items:
        return known_items[template_item]

    # Search within section row range
    if section in section_ranges:
        start_row, end_row = section_ranges[section]
        row = _find_row_in_range(ws, template_item, start_row, end_row)
        if row is not None:
            return row

    # Fallback: search entire column B
    row = _find_row_in_range(ws, template_item, 1, ws.max_row)
    return row


def _parse_numeric_value(value_str: str) -> Optional[float]:
    """
    Parse a numeric value from a string.
    Handles Indian number formatting, parentheses for negatives, etc.

    Examples:
        "1,23,456.78" -> 123456.78
        "(5,432.10)" -> -5432.10
        "−1,000" -> -1000
        "Nil" -> 0.0
        "-" -> 0.0
    """
    if not value_str:
        return None

    s = value_str.strip()

    # Handle special cases
    if s.lower() in ("nil", "na", "n.a.", "n/a", "-", "—", "–"):
        return 0.0

    # Check for negative indicators
    is_negative = False
    if s.startswith("(") and s.endswith(")"):
        is_negative = True
        s = s[1:-1].strip()
    elif s.startswith("−") or s.startswith("–"):
        is_negative = True
        s = s[1:].strip()
    elif s.startswith("-"):
        is_negative = True
        s = s[1:].strip()

    # Remove commas and spaces
    s = s.replace(",", "").replace(" ", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")

    # Remove currency symbols and non-numeric chars (except decimal point)
    s = re.sub(r'[^\d.]', '', s)

    if not s:
        return None

    try:
        value = float(s)
        return -value if is_negative else value
    except ValueError:
        return None


def _apply_formulas(ws, formulas: dict):
    """
    Apply Excel formulas to calculated rows.

    This overwrites any extracted values in formula rows with Excel formulas
    that reference the component cells. This ensures totals auto-compute
    from their components and stay consistent.

    Args:
        ws: The worksheet to modify.
        formulas: Dict mapping row_number -> formula_string.
    """
    for row_num, formula in formulas.items():
        ws.cell(row=row_num, column=VALUES_COLUMN, value=formula)
        logger.debug(f"  Formula at row {row_num}: {formula}")


# ============================================================================
# WRITE FUNCTIONS
# ============================================================================


def write_balance_sheet(
    template_path: str,
    output_path: str,
    mapping_results: list[MappingResult],
    year: int,
) -> dict:
    """
    Write Balance Sheet mapping results to the Excel template.

    All values are written to column C (the "values" column).
    After writing extracted values, Excel formulas are applied to
    calculated/highlighted rows (totals, subtotals).
    """
    wb = openpyxl.load_workbook(template_path)

    if "Balance Sheet" not in wb.sheetnames:
        logger.error(f"Sheet 'Balance Sheet' not found. Available: {wb.sheetnames}")
        return {"written": 0, "skipped": 0, "errors": len(mapping_results)}

    ws = wb["Balance Sheet"]
    stats = {"written": 0, "skipped": 0, "errors": 0}

    for result in mapping_results:
        # Find the row for this item
        row_num = _find_row_for_item(
            ws, result.template_item, result.section,
            BS_SECTION_ROW_RANGES, BS_KNOWN_ITEM_ROWS
        )

        if row_num is None:
            # Try with cleaned label (remove prefix)
            clean_label = re.sub(r'^\(?[a-z]+[\).]\s*', '', result.template_item)
            row_num = _find_row_for_item(
                ws, clean_label, result.section,
                BS_SECTION_ROW_RANGES, BS_KNOWN_ITEM_ROWS
            )

        if row_num is None:
            logger.warning(
                f"BS: Could not find row for '{result.template_item}' "
                f"(section='{result.section}'). Value '{result.value}' skipped."
            )
            stats["skipped"] += 1
            continue

        # Write the value to column C
        try:
            numeric_value = _parse_numeric_value(result.value)
            ws.cell(row=row_num, column=VALUES_COLUMN, value=numeric_value)

            if numeric_value is not None:
                ws.cell(row=row_num, column=VALUES_COLUMN).number_format = '#,##0.00'

            logger.debug(
                f"BS: Wrote '{result.template_item}' = {numeric_value} "
                f"at row {row_num}, col C"
            )
            stats["written"] += 1

        except Exception as e:
            logger.error(f"BS: Error writing '{result.template_item}': {e}")
            stats["errors"] += 1

    # Apply formulas for calculated/highlighted rows
    logger.info(f"BS: Applying {len(BS_FORMULAS)} formulas for calculated fields")
    _apply_formulas(ws, BS_FORMULAS)

    # Save
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    logger.info(f"Saved Balance Sheet to {output_path}")
    return stats


def write_pl_or_cf_sheet(
    file_path: str,
    output_path: str,
    sheet_name: str,
    mapping_results: list[MappingResult],
    year: int,
) -> dict:
    """
    Write P&L or Cash Flow mapping results to the Excel template.

    All values are written to column C (the "values" column).
    For P&L, Excel formulas are applied to calculated/highlighted rows
    after writing extracted values.
    """
    wb = openpyxl.load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return {"written": 0, "skipped": 0, "errors": len(mapping_results)}

    ws = wb[sheet_name]
    stats = {"written": 0, "skipped": 0, "errors": 0}

    # Select section ranges and known items based on sheet type
    if sheet_name == "P&L":
        section_ranges = PL_SECTION_ROW_RANGES
        known_items = PL_KNOWN_ITEM_ROWS
    elif sheet_name == "Cash Flow":
        section_ranges = CF_SECTION_ROW_RANGES
        known_items = {}
    else:
        section_ranges = {}
        known_items = {}

    for result in mapping_results:
        # Find the row for this item
        row_num = _find_row_for_item(
            ws, result.template_item, result.section,
            section_ranges, known_items
        )

        if row_num is None:
            # Try with cleaned label (remove prefix)
            clean_label = re.sub(r'^\(?[a-z]+[\).]\s*', '', result.template_item)
            clean_label = re.sub(r'^[IVX]+\.\s*', '', clean_label)
            row_num = _find_row_for_item(
                ws, clean_label, result.section,
                section_ranges, known_items
            )

        if row_num is None:
            logger.warning(
                f"{sheet_name}: Could not find row for '{result.template_item}' "
                f"(section='{result.section}'). Value '{result.value}' skipped."
            )
            stats["skipped"] += 1
            continue

        # Write the value to column C
        try:
            numeric_value = _parse_numeric_value(result.value)
            ws.cell(row=row_num, column=VALUES_COLUMN, value=numeric_value)

            if numeric_value is not None:
                ws.cell(row=row_num, column=VALUES_COLUMN).number_format = '#,##0.00'

            logger.debug(
                f"{sheet_name}: Wrote '{result.template_item}' = {numeric_value} "
                f"at row {row_num}, col C"
            )
            stats["written"] += 1

        except Exception as e:
            logger.error(f"{sheet_name}: Error writing '{result.template_item}': {e}")
            stats["errors"] += 1

    # Apply formulas for P&L calculated fields
    if sheet_name == "P&L":
        logger.info(f"P&L: Applying {len(PL_FORMULAS)} formulas for calculated fields")
        _apply_formulas(ws, PL_FORMULAS)

    # NOTE: Default zeros for Cash Flow REMOVED per policy — only write 0
    # if the PDF explicitly says 0/Nil/blank/—. Absent items are left blank.

    wb.save(output_path)
    logger.info(f"Saved {sheet_name} to {output_path}")
    return stats


def write_metadata_sheet(
    output_path: str,
    metadata: dict,
) -> dict:
    """
    Write a 'Meta Data' sheet to the output Excel file with extraction metadata.

    The metadata sheet contains key information about the extraction:
        - Company name (from PDF filename)
        - Financial year
        - Value unit (₹ in Lakhs, ₹ in Crores, etc.)
        - Statement type (Standalone / Consolidated)
        - Pages used for each statement
        - Extraction timestamp
        - Completeness percentages
        - Number of rows extracted / mapped / written per sheet

    Args:
        output_path: Path to the existing output Excel file (will be modified).
        metadata: Dict with metadata fields. Expected keys:
            - year: int
            - unit: str (e.g., "₹ in Lakhs")
            - statement_type: str (e.g., "Standalone", "Consolidated")
            - bs_pages: list[int]
            - pl_pages: list[int]
            - cf_pages: list[int]
            - bs_rows_extracted: int
            - pl_rows_extracted: int
            - cf_rows_extracted: int
            - bs_mapped: int
            - pl_mapped: int
            - cf_mapped: int
            - bs_written: int
            - pl_written: int
            - cf_written: int
            - bs_completeness: float
            - pl_completeness: float
            - cf_completeness: float
            - overall_completeness: float
            - pdf_filename: str
            - extraction_timestamp: str

    Returns:
        Dict with write statistics.
    """
    wb = openpyxl.load_workbook(output_path)

    # Remove existing Meta Data sheet if present
    if "Meta Data" in wb.sheetnames:
        del wb["Meta Data"]

    ws = wb.create_sheet("Meta Data", 0)  # Insert as first sheet

    # Styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11, color='1E3A5F')
    section_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    label_font = Font(name='Calibri', size=11, bold=True, color='475569')
    value_font = Font(name='Calibri', size=11, color='1E293B')
    value_font_blue = Font(name='Calibri', size=11, color='2563EB', bold=True)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50

    row = 1

    # --- Title ---
    ws.cell(row=row, column=1, value="EWS Extraction — Meta Data").font = Font(
        name='Calibri', bold=True, size=14, color='1E3A5F'
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    # --- Section: General Information ---
    row = _write_metadata_section_header(ws, row, "General Information", header_font, header_fill, thin_border)

    general_fields = [
        ("PDF Filename", metadata.get("pdf_filename", "N/A")),
        ("Financial Year", str(metadata.get("year", "N/A"))),
        ("Value Unit", metadata.get("unit", "Not specified")),
        ("Statement Type", metadata.get("statement_type", "Not specified")),
        ("Extraction Timestamp", metadata.get("extraction_timestamp", "N/A")),
    ]
    row = _write_metadata_fields(ws, row, general_fields, label_font, value_font, thin_border,
                                  highlight_key="Value Unit", highlight_font=value_font_blue)

    row += 1  # Blank row

    # --- Section: Page Detection ---
    row = _write_metadata_section_header(ws, row, "Page Detection", header_font, header_fill, thin_border)

    bs_pages = metadata.get("bs_pages", [])
    pl_pages = metadata.get("pl_pages", [])
    cf_pages = metadata.get("cf_pages", [])

    page_fields = [
        ("Balance Sheet Pages", ", ".join(str(p) for p in bs_pages) if bs_pages else "Not found"),
        ("P&L Pages", ", ".join(str(p) for p in pl_pages) if pl_pages else "Not found"),
        ("Cash Flow Pages", ", ".join(str(p) for p in cf_pages) if cf_pages else "Not found"),
    ]
    row = _write_metadata_fields(ws, row, page_fields, label_font, value_font, thin_border)

    row += 1  # Blank row

    # --- Section: Extraction Statistics ---
    row = _write_metadata_section_header(ws, row, "Extraction Statistics", header_font, header_fill, thin_border)

    stats_fields = [
        ("Balance Sheet — Rows Extracted", str(metadata.get("bs_rows_extracted", 0))),
        ("Balance Sheet — Items Mapped", str(metadata.get("bs_mapped", 0))),
        ("Balance Sheet — Values Written", str(metadata.get("bs_written", 0))),
        ("P&L — Rows Extracted", str(metadata.get("pl_rows_extracted", 0))),
        ("P&L — Items Mapped", str(metadata.get("pl_mapped", 0))),
        ("P&L — Values Written", str(metadata.get("pl_written", 0))),
        ("Cash Flow — Rows Extracted", str(metadata.get("cf_rows_extracted", 0))),
        ("Cash Flow — Items Mapped", str(metadata.get("cf_mapped", 0))),
        ("Cash Flow — Values Written", str(metadata.get("cf_written", 0))),
    ]
    row = _write_metadata_fields(ws, row, stats_fields, label_font, value_font, thin_border)

    row += 1  # Blank row

    # --- Section: Completeness ---
    row = _write_metadata_section_header(ws, row, "Completeness", header_font, header_fill, thin_border)

    bs_comp = metadata.get("bs_completeness", 0)
    pl_comp = metadata.get("pl_completeness", 0)
    cf_comp = metadata.get("cf_completeness", 0)
    overall_comp = metadata.get("overall_completeness", 0)

    completeness_fields = [
        ("Balance Sheet Completeness", f"{bs_comp:.1%}"),
        ("P&L Completeness", f"{pl_comp:.1%}"),
        ("Cash Flow Completeness", f"{cf_comp:.1%}"),
        ("Overall Completeness", f"{overall_comp:.1%}"),
    ]
    row = _write_metadata_fields(ws, row, completeness_fields, label_font, value_font, thin_border,
                                  highlight_key="Overall Completeness", highlight_font=value_font_blue)

    row += 1  # Blank row

    # --- Section: CA Validation ---
    ca_flags_count = metadata.get("ca_flags_count", 0)
    if ca_flags_count > 0 or metadata.get("ca_inferred", 0) > 0:
        row = _write_metadata_section_header(ws, row, "CA Validation & Inference", header_font, header_fill, thin_border)

        ca_balanced = metadata.get("ca_balanced", True)
        ca_fields = [
            ("BS Equation Balanced", "YES" if ca_balanced else "NO — check mapping"),
            ("CA Validation Errors", str(metadata.get("ca_errors", 0))),
            ("CA Validation Warnings", str(metadata.get("ca_warnings", 0))),
            ("CA Inferred Mappings Applied", str(metadata.get("ca_inferred", 0))),
            ("Items with Note References", str(metadata.get("ca_note_refs", 0))),
            ("Total CA Flags", str(ca_flags_count)),
        ]
        row = _write_metadata_fields(ws, row, ca_fields, label_font, value_font, thin_border,
                                      highlight_key="BS Equation Balanced",
                                      highlight_font=Font(name='Calibri', size=11, color='059669', bold=True)
                                      if ca_balanced
                                      else Font(name='Calibri', size=11, color='DC2626', bold=True))

        # Write CA flag details (up to 20 most important)
        ca_flags_detail = metadata.get("ca_flags_detail", [])
        if ca_flags_detail:
            row += 1
            cell_a = ws.cell(row=row, column=1, value="Validation Flags Detail:")
            cell_a.font = Font(name='Calibri', size=10, bold=True, color='475569')
            row += 1

            # Sort: errors first, then warnings, then info
            severity_order = {"error": 0, "warning": 1, "info": 2}
            ca_flags_detail_sorted = sorted(
                ca_flags_detail,
                key=lambda f: severity_order.get(f.get("severity", "info"), 3)
            )

            for flag in ca_flags_detail_sorted[:20]:
                severity = flag.get("severity", "info").upper()
                category = flag.get("category", "")
                message = flag.get("message", "")
                item = flag.get("item_name", "")
                severity_colors = {
                    "ERROR": "DC2626",
                    "WARNING": "D97706",
                    "INFO": "2563EB",
                }
                color = severity_colors.get(severity, "6B7280")
                flag_text = f"[{severity}] [{category}] {message}"
                if item:
                    flag_text += f" (item: {item})"
                cell = ws.cell(row=row, column=1, value=flag_text)
                cell.font = Font(name='Calibri', size=9, color=color)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                row += 1

        # Write CA inferred mapping details
        ca_inferred_detail = metadata.get("ca_inferred_detail", [])
        if ca_inferred_detail:
            row += 1
            cell_a = ws.cell(row=row, column=1, value="CA Inferred Mappings Detail:")
            cell_a.font = Font(name='Calibri', size=10, bold=True, color='475569')
            row += 1

            for im in ca_inferred_detail:
                item = im.get("template_item", "")
                value = im.get("value", "")
                method = im.get("method", "")
                confidence = im.get("confidence", 0)
                reasoning = im.get("reasoning", "")
                im_text = f"{item} = {value} (method={method}, confidence={confidence:.0%})"
                cell = ws.cell(row=row, column=1, value=im_text)
                cell.font = Font(name='Calibri', size=9, color='059669')
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                row += 1
                if reasoning:
                    cell_r = ws.cell(row=row, column=1, value=f"  → {reasoning[:120]}")
                    cell_r.font = Font(name='Calibri', size=9, color='6B7280', italic=True)
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    row += 1

    # Save
    wb.save(output_path)
    logger.info(f"Saved Meta Data sheet to {output_path}")
    return {"written": 1}


def _write_metadata_section_header(ws, row, title, header_font, header_fill, border):
    """Write a section header row in the Meta Data sheet."""
    cell_a = ws.cell(row=row, column=1, value=title)
    cell_a.font = header_font
    cell_a.fill = header_fill
    cell_a.border = border
    cell_a.alignment = Alignment(horizontal='left', vertical='center')

    cell_b = ws.cell(row=row, column=2, value="")
    cell_b.fill = header_fill
    cell_b.border = border

    row += 1
    return row


def _write_metadata_fields(ws, row, fields, label_font, value_font, border,
                            highlight_key=None, highlight_font=None):
    """
    Write label-value field rows in the Meta Data sheet.

    Args:
        ws: Worksheet.
        row: Current row number.
        fields: List of (label, value) tuples.
        label_font: Font for labels.
        value_font: Font for values.
        border: Cell border style.
        highlight_key: If a field's label matches this, use highlight_font instead.
        highlight_font: Font to use for highlighted values.
    """
    for label, value in fields:
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = label_font
        cell_a.border = border
        cell_a.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        cell_b = ws.cell(row=row, column=2, value=value)
        if highlight_key and label == highlight_key and highlight_font:
            cell_b.font = highlight_font
        else:
            cell_b.font = value_font
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal='left', vertical='center')

        row += 1

    return row


def write_raw_table_sheet(
    output_path: str,
    sheet_name: str,
    extracted_table: ExtractedTable,
) -> dict:
    """
    Write a raw extracted table as a separate sheet in the output Excel file.

    This provides transparency by showing the raw data extracted from the PDF
    before any mapping/filtering was applied. Users can verify the source data
    and compare it against the mapped values in the main sheets.

    The sheet contains columns:
        - Row #: Sequential row number
        - Label: The row label as extracted from the PDF
        - Note: Note reference (e.g., "2", "3A")
        - Year 1: First year value (typically current year)
        - Year 2: Second year value (typically previous year)
        - Indent: Indent level (0=top, 1=sub-item, 2=sub-sub-item)
        - Section Header: "Yes" if this row is a section header
        - Total Row: "Yes" if this row is a total/subtotal
        - Raw Text: The original raw text line from the PDF

    Args:
        output_path: Path to the existing output Excel file (will be modified).
        sheet_name: Name for the new sheet (e.g., "Raw BS", "Raw P&L", "Raw CF").
        extracted_table: ExtractedTable object with raw extracted data.

    Returns:
        Dict with write statistics.
    """
    if not extracted_table or not extracted_table.rows:
        logger.info(f"No raw data to write for '{sheet_name}'")
        return {"written": 0, "note": "no data"}

    wb = openpyxl.load_workbook(output_path)

    # Remove existing sheet if present
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Styles
    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Calibri', size=10)
    cell_font_indent = Font(name='Calibri', size=10, color='6B7280')
    section_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    total_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # Determine number of value columns
    max_values = 0
    for row in extracted_table.rows:
        if len(row.values) > max_values:
            max_values = len(row.values)

    # Build headers
    headers = ["Row #", "Label", "Note"]
    for i in range(max_values):
        if extracted_table.year_headers and i < len(extracted_table.year_headers):
            headers.append(extracted_table.year_headers[i])
        elif i == 0:
            headers.append("Current Year")
        elif i == 1:
            headers.append("Previous Year")
        else:
            headers.append(f"Year {i+1}")
    headers.extend(["Indent", "Section Header", "Total Row", "Raw Text"])

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 6    # Row #
    ws.column_dimensions['B'].width = 50    # Label
    ws.column_dimensions['C'].width = 8     # Note
    for i in range(max_values):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 18
    indent_col = get_column_letter(4 + max_values)
    ws.column_dimensions[indent_col].width = 8
    sec_col = get_column_letter(5 + max_values)
    ws.column_dimensions[sec_col].width = 14
    total_col = get_column_letter(6 + max_values)
    ws.column_dimensions[total_col].width = 10
    raw_col = get_column_letter(7 + max_values)
    ws.column_dimensions[raw_col].width = 60

    # Write data rows
    written = 0
    for row_idx, row in enumerate(extracted_table.rows, 2):
        # Row number
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border

        # Label (with indent visualization)
        label = row.label or ""
        if row.indent_level and row.indent_level > 0:
            indent_prefix = "  " * row.indent_level
            label_cell = ws.cell(row=row_idx, column=2, value=f"{indent_prefix}{label}")
            label_cell.font = cell_font_indent
        else:
            label_cell = ws.cell(row=row_idx, column=2, value=label)
            label_cell.font = cell_font
        label_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Note reference
        ws.cell(row=row_idx, column=3, value=row.note_ref or "").border = thin_border

        # Values
        for i in range(max_values):
            val = row.values[i] if i < len(row.values) else ""
            val_cell = ws.cell(row=row_idx, column=4 + i, value=val)
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal='right', vertical='center')
            val_cell.font = cell_font

        # Indent level
        ws.cell(row=row_idx, column=4 + max_values, value=row.indent_level or 0).border = thin_border

        # Section header flag
        sec_val = "Yes" if row.is_section_header else ""
        ws.cell(row=row_idx, column=5 + max_values, value=sec_val).border = thin_border

        # Total row flag
        total_val = "Yes" if row.is_total else ""
        ws.cell(row=row_idx, column=6 + max_values, value=total_val).border = thin_border

        # Raw text
        raw_text = row.raw_text or ""
        ws.cell(row=row_idx, column=7 + max_values, value=raw_text).border = thin_border
        ws.cell(row=row_idx, column=7 + max_values).font = Font(name='Calibri', size=9, color='9CA3AF')

        # Apply row highlighting for section headers and totals
        if row.is_section_header:
            for col in range(1, 8 + max_values):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = section_fill
                if col == 2:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='374151')
        elif row.is_total:
            for col in range(1, 8 + max_values):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = total_fill
                if col == 2:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='92400E')

        written += 1

    # Add page numbers info at the bottom
    info_row = len(extracted_table.rows) + 3
    if extracted_table.page_numbers:
        ws.cell(row=info_row, column=1, value="Source Pages:").font = Font(
            name='Calibri', size=9, bold=True, color='6B7280'
        )
        ws.cell(row=info_row, column=2, value=", ".join(str(p) for p in extracted_table.page_numbers)).font = Font(
            name='Calibri', size=9, color='6B7280'
        )

    if extracted_table.unit and extracted_table.unit != "Not specified":
        ws.cell(row=info_row + 1, column=1, value="Unit:").font = Font(
            name='Calibri', size=9, bold=True, color='6B7280'
        )
        ws.cell(row=info_row + 1, column=2, value=extracted_table.unit).font = Font(
            name='Calibri', size=9, color='6B7280'
        )

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(7 + max_values)}{len(extracted_table.rows) + 1}"

    wb.save(output_path)
    logger.info(f"Saved raw table sheet '{sheet_name}' with {written} rows to {output_path}")
    return {"written": written}


def write_all_sheets(
    template_path: str,
    output_path: str,
    bs_results: list[MappingResult],
    pl_results: list[MappingResult],
    cf_results: list[MappingResult],
    year: int,
    metadata: dict = None,
    bs_table: ExtractedTable = None,
    pl_table: ExtractedTable = None,
    cf_table: ExtractedTable = None,
) -> dict:
    """
    Write mapping results for all three financial statement sheets.

    Also writes raw extracted table sheets for transparency, showing the
    original data from the PDF before any mapping/filtering was applied.

    Args:
        template_path: Path to the template Excel file.
        output_path: Path to write the output Excel file.
        bs_results: Balance Sheet mapping results.
        pl_results: P&L mapping results.
        cf_results: Cash Flow mapping results.
        year: The financial year.
        metadata: Optional dict with extraction metadata for the Meta Data sheet.
        bs_table: Optional ExtractedTable with raw BS data for the "Raw BS" sheet.
        pl_table: Optional ExtractedTable with raw P&L data for the "Raw P&L" sheet.
        cf_table: Optional ExtractedTable with raw CF data for the "Raw CF" sheet.

    Returns:
        Dict with per-sheet statistics.
    """
    stats = {}

    # Step 1: Write Balance Sheet (creates output file from template)
    if bs_results:
        stats["balance_sheet"] = write_balance_sheet(
            template_path, output_path, bs_results, year
        )
    else:
        # Just copy the template to output
        import shutil
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        shutil.copy2(template_path, output_path)
        stats["balance_sheet"] = {"written": 0, "skipped": 0, "errors": 0, "note": "no data"}

    # Step 2: Write P&L (modifies the output file)
    if pl_results:
        stats["profit_and_loss"] = write_pl_or_cf_sheet(
            output_path, output_path, "P&L", pl_results, year
        )
    else:
        stats["profit_and_loss"] = {"written": 0, "skipped": 0, "errors": 0, "note": "no data"}

    # Step 3: Write Cash Flow (modifies the output file)
    if cf_results:
        stats["cash_flow"] = write_pl_or_cf_sheet(
            output_path, output_path, "Cash Flow", cf_results, year
        )
    else:
        stats["cash_flow"] = {"written": 0, "skipped": 0, "errors": 0, "note": "no data"}

    # Step 4: Write raw extracted table sheets (for transparency)
    if bs_table and bs_table.rows:
        stats["raw_bs"] = write_raw_table_sheet(output_path, "Raw BS", bs_table)
    if pl_table and pl_table.rows:
        stats["raw_pl"] = write_raw_table_sheet(output_path, "Raw P&L", pl_table)
    if cf_table and cf_table.rows:
        stats["raw_cf"] = write_raw_table_sheet(output_path, "Raw CF", cf_table)

    # Step 5: Write Meta Data sheet (if metadata provided)
    if metadata:
        stats["metadata"] = write_metadata_sheet(output_path, metadata)

    return stats
