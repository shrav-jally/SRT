"""
CA RAG Agent — "Extra CA" for Validation & Filling

This agent acts as an additional Chartered Accountant who:
1. Reads the extraction results from the main pipeline (Excel)
2. For each template item, generates targeted questions for RAG
3. Validates prefilled values against RAG-retrieved context
4. Fills previously empty cells using RAG-retrieved context
5. Generates a new Excel with color-coded changes:
   - YELLOW: RAG verified and CHANGED a value (different from original)
   - GREEN:  RAG filled a previously EMPTY cell (new data)
   - No color: RAG verified and value matches (no change needed)

Works for STANDALONE financial statements only (per current policy).

The prompts are auto-generated from the extraction template items,
so the agent knows exactly what to ask RAG for each line item.
"""

import json
import logging
import os
import re
from dataclasses import dataclass
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from . import config
from . import vector_store

logger = logging.getLogger(__name__)


# ============================================================================
# COLOR DEFINITIONS FOR EXCEL
# ============================================================================

# Yellow: RAG verified and CHANGED a value (original had different value)
FILL_CHANGED = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Green: RAG filled a previously EMPTY cell (original was blank)
FILL_NEW = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# Light blue: RAG verified and value MATCHES (no change, just confirmation)
FILL_VERIFIED = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Font for changed/new cells (bold to highlight)
FONT_HIGHLIGHT = Font(bold=True)
FONT_NORMAL = Font()


# ============================================================================
# TEMPLATE ITEMS FOR PROMPT GENERATION
# ============================================================================

# These are the SAME template items from ews_agent/data_mapper.py
# Duplicated here to keep rag_agent isolated from ews_agent.

BALANCE_SHEET_ITEMS = {
    "Non-current assets": [
        "(a) Property, Plant and Equipment",
        "(b) Capital work-in-progress",
        "(c) Investment Property",
        "(d) Goodwill",
        "(e) Other Intangible assets",
        "(f) Intangible assets under development",
        "(g) Biological Assets other than bearer plants",
        "(h) Financial Assets",
        "(i) Investments",
        "(ii) Trade receivables",
        "(iii) Loans",
        "(iv) Others (to be specified)",
        "(i) Deferred tax assets (net)",
        "(j) Other non-current assets",
    ],
    "Current assets": [
        "(a) Inventories",
        "(b) Financial Assets",
        "(i) Investments",
        "(ii) Trade receivables",
        "(iii) Cash and cash equivalents",
        "(iv) Bank balances other than (iii) above",
        "(v) Loans",
        "(vi) Others (to be specified)",
        "(c) Current Tax Assets (Net)",
        "(d) Other current assets",
    ],
    "Equity": [
        "Equity Share Capital",
        "Other Equity",
    ],
    "Non-current liabilities": [
        "(a) Financial Liabilities",
        "(i) Borrowings",
        "(ii) Trade Payables",
        "(A) total outstanding dues of micro enterprises and small enterprises; and",
        "(B) total outstanding dues of creditors other than micro enterprises and small enterprises.",
        "(iii) Other financial liabilities",
        "(b) Provisions",
        "(c) Deferred tax liabilities (Net)",
        "(d) Other non-current liabilities",
    ],
    "Current liabilities": [
        "(a) Financial Liabilities",
        "(i) Borrowings",
        "(ii) Trade payables",
        "(iii) Other financial liabilities",
        "(b) Other current liabilities",
        "(c) Provisions",
        "(d) Current Tax Liabilities (Net)",
    ],
}

PL_ITEMS = {
    "Income": [
        "I. Revenue from operations",
        "II. Other income",
    ],
    "IV. Expenses": [
        "Cost of materials consumed",
        "Purchases of Stock-in-Trade",
        "Changes in inventories of finished goods, work-in-progress and Stock-in-Trade",
        "Employee benefits expense",
        "Finance costs",
        "Depreciation and amortisation expense",
        "Other expenses",
    ],
    "Taxes": [
        "(1) Current tax",
        "(2) Deferred tax",
    ],
}

CF_ITEMS = {
    "Cash Flow": [
        "Cash Flows From Operations",
        "Cash Flows From Investing",
        "Cash Flows From Financing",
    ],
    "Other Financial Information": [
        "Contingent Liabilities",
        "Creditors outstanding for more than 1 year",
        "Debtors outstanding for more than 1 year",
        "Inventory outstanding for more than 180 days",
        "Disputed Trade Receivable",
        "Advance from Customers",
        "Power and fuel/electricity Expenses",
        "Significant impairment of assets/ write-offs",
        "Auditors Remunerations",
        "One-time revenue (revaluation of assets, etc.)",
        "Provision for doubtful debts expense",
        "Bad Debts expenses",
        "RP Investments",
        "RP Expenses",
        "RP Revenues",
        "RP Loan and Advances*",
        "RP Bad debts",
        "RP Loan (Liab)",
        "Current maturities of borrowings/debts, including interest",
    ],
}


# ============================================================================
# PROMPT GENERATION
# ============================================================================


# Classification of how a CA would determine each template item's value
# - "direct": Value is explicitly stated in the financial statement
# - "calculated": Value is a sum/subtotal of sub-items (e.g., Total Assets = Non-current + Current)
# - "inferred": Value can be derived from accounting equations (e.g., Other Equity = Total Equity - Share Capital)
VALUE_SOURCE_CLASSIFICATION = {
    "balance_sheet": {
        "direct": [
            "(a) Property, Plant and Equipment", "(b) Capital work-in-progress",
            "(c) Investment Property", "(d) Goodwill", "(e) Other Intangible assets",
            "(f) Intangible assets under development",
            "(g) Biological Assets other than bearer plants",
            "(i) Investments", "(ii) Trade receivables", "(iii) Loans",
            "(iv) Others (to be specified)", "(i) Deferred tax assets (net)",
            "(j) Other non-current assets",
            "(a) Inventories", "(i) Investments", "(ii) Trade receivables",
            "(iii) Cash and cash equivalents", "(iv) Bank balances other than (iii) above",
            "(v) Loans", "(vi) Others (to be specified)",
            "(c) Current Tax Assets (Net)", "(d) Other current assets",
            "Equity Share Capital", "Other Equity",
            "(i) Borrowings", "(ii) Trade Payables",
            "(A) total outstanding dues of micro enterprises and small enterprises; and",
            "(B) total outstanding dues of creditors other than micro enterprises and small enterprises.",
            "(iii) Other financial liabilities", "(b) Provisions",
            "(c) Deferred tax liabilities (Net)", "(d) Other non-current liabilities",
            "(i) Borrowings", "(ii) Trade payables",
            "(iii) Other financial liabilities", "(b) Other current liabilities",
            "(c) Provisions", "(d) Current Tax Liabilities (Net)",
        ],
        "calculated": [
            "(h) Financial Assets",  # Sum of sub-items (i)+(ii)+(iii)+(iv)
            "(b) Financial Assets",  # Sum of sub-items (i)+(ii)+(iii)+(iv)+(v)+(vi)
            "(a) Financial Liabilities",  # Sum of sub-items
        ],
        "inferred": [
            # These can be derived from BS equation: Assets = Equity + Liabilities
            # Or from sub-section totals
        ],
    },
    "profit_and_loss": {
        "direct": [
            "I. Revenue from operations", "II. Other income",
            "Cost of materials consumed", "Purchases of Stock-in-Trade",
            "Changes in inventories of finished goods, work-in-progress and Stock-in-Trade",
            "Employee benefits expense", "Finance costs",
            "Depreciation and amortisation expense", "Other expenses",
            "(1) Current tax", "(2) Deferred tax",
        ],
        "calculated": [],  # P&L totals are formula items (skipped)
        "inferred": [],
    },
    "cash_flow": {
        "direct": [
            "Cash Flows From Operations", "Cash Flows From Investing",
            "Cash Flows From Financing",
            "Contingent Liabilities", "Creditors outstanding for more than 1 year",
            "Debtors outstanding for more than 1 year",
            "Inventory outstanding for more than 180 days",
            "Disputed Trade Receivable", "Advance from Customers",
            "Power and fuel/electricity Expenses",
            "Significant impairment of assets/ write-offs",
            "Auditors Remunerations",
            "One-time revenue (revaluation of assets, etc.)",
            "Provision for doubtful debts expense", "Bad Debts expenses",
            "RP Investments", "RP Expenses", "RP Revenues",
            "RP Loan and Advances*", "RP Bad debts", "RP Loan (Liab)",
            "Current maturities of borrowings/debts, including interest",
        ],
        "calculated": [],
        "inferred": [],
    },
}


def _get_value_source(statement_type: str, item: str) -> str:
    """Classify how a CA would determine this item's value."""
    classification = VALUE_SOURCE_CLASSIFICATION.get(statement_type, {})
    if item in classification.get("direct", []):
        return "Direct — value explicitly stated in the statement"
    elif item in classification.get("calculated", []):
        return "Calculated — sum of sub-items (CA would add up components)"
    elif item in classification.get("inferred", []):
        return "Inferred — derived from accounting equations (e.g., BS balancing)"
    else:
        return "Direct or Inferred — look for explicit value first; if not found, a CA would infer from context"


def _generate_rag_questions(statement_type: str, items_dict: dict,
                            financial_year: str = None) -> list[dict]:
    """
    Generate targeted RAG questions for each template item.

    For each item, we create a specific question that asks RAG to find
    the value for that line item from the standalone financial statements.
    Questions are framed from a CA's perspective — the value may be:
    1. DIRECTLY PRESENT: Explicitly stated in the financial statement
    2. CALCULATED: A sum/subtotal of sub-items (e.g., Total Non-Current Assets)
    3. INFERRED: Derived from accounting equations (e.g., BS equation, P&L waterfall)

    Args:
        statement_type: "balance_sheet", "profit_and_loss", or "cash_flow"
        items_dict: Dict of section -> list of item names
        financial_year: Financial year string (e.g., "2018", "2017-18") from Meta Data.
                        Included in questions so RAG targets the correct year column.

    Returns:
        List of dicts with 'item', 'section', 'question', 'value_source' keys.
    """
    statement_name = {
        "balance_sheet": "Balance Sheet",
        "profit_and_loss": "Statement of Profit and Loss",
        "cash_flow": "Cash Flow Statement",
    }.get(statement_type, statement_type)

    # Build year context string for questions
    year_context = ""
    if financial_year and financial_year not in ("N/A", "None", ""):
        # Handle both "2018" and "2017-18" formats
        if "-" in str(financial_year):
            year_context = f" for FY {financial_year}"
        else:
            # Single year — derive FY range (e.g., 2018 → FY 2017-18)
            yr = int(financial_year) if str(financial_year).isdigit() else financial_year
            if isinstance(yr, int):
                year_context = f" for FY {yr-1}-{str(yr)[2:]}"
            else:
                year_context = f" for the year {financial_year}"

    # Build year column instruction
    year_col_instruction = ""
    if financial_year and financial_year not in ("N/A", "None", ""):
        if "-" in str(financial_year):
            # FY 2017-18 → year ending March 31, 2018
            end_year = str(financial_year).split("-")[-1]
            if len(end_year) == 2:
                end_year = "20" + end_year
            year_col_instruction = (
                f"Extract the value from the column for the year ending March 31, {end_year} "
                f"(current year). Do NOT use the previous year column."
            )
        elif str(financial_year).isdigit():
            year_col_instruction = (
                f"Extract the value from the column for the year ending March 31, {financial_year} "
                f"(current year). Do NOT use the previous year column."
            )

    questions = []

    for section, items in items_dict.items():
        for item in items:
            # Clean item name for the question
            clean_item = re.sub(r'^\(?[a-zA-Z]+[\).]\s*', '', item).strip()
            clean_item = re.sub(r'^[IVXivx]+\.\s*', '', clean_item).strip()

            # Get value source classification
            value_source = _get_value_source(statement_type, item)

            # Build CA-level instruction based on value source
            if "Calculated" in value_source:
                ca_hint = (
                    "Think like a CA: This is a subtotal/calculated line item. "
                    "If the value is not explicitly stated, sum up its sub-components. "
                    "For example, 'Financial Assets' = Investments + Trade Receivables + Loans + Others."
                )
            elif "Inferred" in value_source:
                ca_hint = (
                    "Think like a CA: This value may need to be inferred from accounting equations. "
                    "For example, if Total Assets = Total Equity + Liabilities, and you know two of three, "
                    "derive the third. Use the balance sheet equation or P&L waterfall logic."
                )
            else:
                ca_hint = (
                    "Think like a CA: Look for the value directly in the statement first. "
                    "If not explicitly stated, check if it can be calculated from sub-items "
                    "or inferred from accounting relationships (e.g., BS equation, P&L totals)."
                )

            # Build year-specific question text
            if statement_type == "balance_sheet":
                question = (
                    f"In the standalone Balance Sheet{year_context}, what is the value of "
                    f'"{clean_item}" under the section "{section}"? '
                    f"Provide the exact numeric value from the current year column. "
                    f"{year_col_instruction} "
                    f"If the value is zero or nil, state '0'. If not found, state 'NOT FOUND'. "
                    f"{ca_hint}"
                )
            elif statement_type == "profit_and_loss":
                question = (
                    f"In the standalone Statement of Profit and Loss{year_context}, what is the value of "
                    f'"{clean_item}" under "{section}"? '
                    f"Provide the exact numeric value from the current year column. "
                    f"{year_col_instruction} "
                    f"If the value is zero or nil, state '0'. If not found, state 'NOT FOUND'. "
                    f"{ca_hint}"
                )
            elif statement_type == "cash_flow":
                if section == "Other Financial Information":
                    question = (
                        f"In the Notes to Accounts of the standalone financial statements{year_context}, "
                        f"what is the value of \"{clean_item}\"? "
                        f"Provide the exact numeric value. "
                        f"{year_col_instruction} "
                        f"If the value is zero or nil, state '0'. If not found, state 'NOT FOUND'. "
                        f"{ca_hint}"
                    )
                else:
                    question = (
                        f"In the standalone Cash Flow Statement{year_context}, what is the net cash flow "
                        f"from {clean_item.replace('Cash Flows From ', '').lower()}? "
                        f"Provide the exact numeric value from the current year column. "
                        f"{year_col_instruction} "
                        f"If not found, state 'NOT FOUND'. "
                        f"{ca_hint}"
                    )
            else:
                question = (
                    f"What is the value of \"{clean_item}\" in the financial statements{year_context}? "
                    f"Provide the exact numeric value. If not found, state 'NOT FOUND'. "
                    f"{ca_hint}"
                )

            questions.append({
                "item": item,
                "section": section,
                "question": question,
                "value_source": value_source,
            })

    return questions


# ============================================================================
# LLM HELPER (uses shared ews_agent.llm_config)
# ============================================================================

_llm_instance = None


def _get_llm():
    """Get or create the LLM instance for CA RAG agent (uses shared llm_config)."""
    global _llm_instance
    if _llm_instance is not None:
        return _llm_instance

    try:
        from ews_agent.llm_config import get_llm

        # Low temperature + low max_tokens: CA agent only needs short numeric answers
        # Model has 20K context window; 500 output tokens is plenty for "6,527.97" style answers
        _llm_instance = get_llm(temperature=0.05, max_tokens=500)
        logger.info("CA RAG Agent LLM initialized via ews_agent.llm_config.get_llm()")
        return _llm_instance
    except Exception as e:
        logger.error(f"Failed to initialize CA RAG LLM: {e}")
        raise


# ============================================================================
# RAG-BASED VALUE EXTRACTION
# ============================================================================


def _extract_value_from_rag(
    question: str,
    filter_filename: str = None,
    top_k: int = None,
    financial_year: str = None,
) -> dict:
    """
    Use RAG to extract a value for a specific template item.

    Steps:
    1. Retrieve top-k relevant chunks from vector store
    2. Build prompt asking LLM to extract the value from context
    3. Parse the LLM response for a numeric value

    Args:
        question: The targeted question for this item.
        filter_filename: If set, only search within this PDF.
        top_k: Number of chunks to retrieve.
        financial_year: Financial year string (e.g., "2018", "2017-18") from Meta Data.
                        Included in LLM prompt so it targets the correct year column.

    Returns:
        Dict with 'value', 'confidence', 'sources', 'raw_answer'.
        value is None if not found.
    """
    if top_k is None:
        top_k = config.RAG_TOP_K_CA

    # Step 1: Retrieve relevant chunks
    search_results = vector_store.search(
        query=question,
        top_k=top_k,
        filter_filename=filter_filename,
    )

    if not search_results:
        return {
            "value": None,
            "confidence": 0.0,
            "sources": [],
            "raw_answer": "No relevant pages found in RAG store",
        }

    # Step 2: Build context
    context_parts = []
    sources = []
    for result in search_results:
        meta = result["metadata"]
        page = meta.get("page_number", "?")
        fname = meta.get("pdf_filename", "unknown")
        context_parts.append(f"[Page {page} of {fname}]:\n{result['text']}")
        sources.append({
            "filename": fname,
            "page": page,
            "distance": round(result["distance"], 4),
        })

    context_text = "\n\n---\n\n".join(context_parts)

    # Build year-specific instruction for the LLM prompt
    year_instruction = ""
    if financial_year and financial_year not in ("N/A", "None", ""):
        if "-" in str(financial_year):
            end_year = str(financial_year).split("-")[-1]
            if len(end_year) == 2:
                end_year = "20" + end_year
            year_instruction = (
                f"\nCRITICAL: This is for FY {financial_year}. "
                f"You MUST extract the value from the column for the year ending March 31, {end_year} "
                f"(current year). Do NOT use the previous year column. "
                f"If the context shows multiple year columns, pick ONLY the {end_year} column."
            )
        elif str(financial_year).isdigit():
            year_instruction = (
                f"\nCRITICAL: This is for FY {int(financial_year)-1}-{str(financial_year)[2:]}. "
                f"You MUST extract the value from the column for the year ending March 31, {financial_year} "
                f"(current year). Do NOT use the previous year column. "
                f"If the context shows multiple year columns, pick ONLY the {financial_year} column."
            )

    # Step 3: Ask LLM to extract the value
    prompt = f"""You are a Chartered Accountant reviewing Indian company annual reports (standalone financial statements).

Based ONLY on the context below, answer the question. Extract the exact numeric value.
{year_instruction}

Rules:
- If the value is found, respond with JUST the number (e.g., "6,527.97" or "(1,234.56)" for negative).
- If the value is explicitly stated as zero, nil, or dash, respond with "0".
- If the value is NOT found in the context, respond with "NOT FOUND".
- Do NOT guess or calculate. Only report values explicitly stated in the text.
- For negative values, use parentheses format: (1,234.56)
- IMPORTANT: If the financial statement shows two year columns, extract ONLY from the CURRENT year column (as specified above), NOT the comparative/previous year.

Context:
{context_text}

Question: {question}

Answer:"""

    try:
        llm = _get_llm()
        from langchain_core.messages import HumanMessage
        response = llm.invoke([HumanMessage(content=prompt)])
        raw_answer = response.content if hasattr(response, 'content') else str(response)
    except Exception as e:
        logger.error(f"CA RAG LLM call failed: {e}")
        return {
            "value": None,
            "confidence": 0.0,
            "sources": sources,
            "raw_answer": f"LLM error: {str(e)}",
        }

    # Step 4: Parse the answer
    raw_answer = raw_answer.strip()

    if raw_answer.upper() in ("NOT FOUND", "N/A", "NA", "-"):
        return {
            "value": None,
            "confidence": 0.0,
            "sources": sources,
            "raw_answer": raw_answer,
        }

    # Try to extract a numeric value from the answer
    value = _parse_numeric_value(raw_answer)

    confidence = 0.8 if value is not None else 0.0
    if value is not None and len(sources) > 0 and sources[0]["distance"] < 0.3:
        confidence = 0.9  # High confidence if very relevant context

    return {
        "value": value,
        "confidence": confidence,
        "sources": sources,
        "raw_answer": raw_answer,
    }


def _parse_numeric_value(text: str) -> Optional[str]:
    """
    Parse a numeric value from LLM response text.

    Handles formats like:
    - "6,527.97"
    - "(1,234.56)"  (negative in Indian accounting)
    - "₹1,234.56"
    - "Rs. 1,234.56"
    - "0"
    - "Nil" → "0"

    Returns:
        String representation of the value, or None if not parseable.
    """
    text = text.strip()

    # Handle explicit zero/nil
    if text.lower() in ("nil", "na", "n.a.", "n/a", "-", "—", "–", "0", "0.00"):
        return "0"

    # Remove currency symbols and prefixes
    text = re.sub(r'[₹$]', '', text)
    text = re.sub(r'Rs\.?\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'INR\s*', '', text, flags=re.IGNORECASE)

    # Check for parentheses (negative value in accounting)
    is_negative = False
    paren_match = re.match(r'^\(([\d,.]+)\)$', text.strip())
    if paren_match:
        is_negative = True
        text = paren_match.group(1)

    # Extract numeric pattern
    num_match = re.search(r'[\d,]+\.?\d*', text)
    if num_match:
        value_str = num_match.group()
        # Validate it looks like a number
        try:
            float(value_str.replace(",", ""))
        except ValueError:
            return None

        if is_negative:
            return f"({value_str})"
        return value_str

    return None


# ============================================================================
# MAIN CA RAG AGENT
# ============================================================================


@dataclass
class RAGValidationResult:
    """Result of CA RAG agent validation for a single item."""
    item: str
    section: str
    original_value: Optional[str]  # Value from extraction pipeline
    rag_value: Optional[str]       # Value from RAG
    action: str                    # "verified", "changed", "filled", "not_found", "skipped"
    confidence: float
    sources: list
    raw_answer: str


# ============================================================================
# PROGRESS TRACKER (for UI progress bars)
# ============================================================================

_progress_store = {}  # session_id -> progress dict


def get_progress(session_id: str) -> dict:
    """Get the current progress for a RAG validation session."""
    return _progress_store.get(session_id, {
        "status": "idle",
        "current_step": "",
        "progress": 0,
        "total_items": 0,
        "completed_items": 0,
        "details": [],
    })


def _update_progress(session_id: str, step: str, progress: float,
                     total: int = 0, completed: int = 0, detail: str = ""):
    """Update progress for a session."""
    if session_id not in _progress_store:
        _progress_store[session_id] = {
            "status": "running",
            "current_step": "",
            "progress": 0,
            "total_items": 0,
            "completed_items": 0,
            "details": [],
        }
    p = _progress_store[session_id]
    p["status"] = "running"
    p["current_step"] = step
    p["progress"] = round(progress, 1)
    if total:
        p["total_items"] = total
    if completed:
        p["completed_items"] = completed
    if detail:
        p["details"].append(detail)
        # Keep only last 50 details
        if len(p["details"]) > 50:
            p["details"] = p["details"][-50:]


def run_ca_rag_validation(
    extraction_excel_path: str,
    pdf_filename: str = None,
    statement_types: list = None,
    session_id: str = "default",
) -> dict:
    """
    Run the CA RAG agent to validate and fill extraction results.

    For each template item in the extraction Excel:
    1. Read the current value from the Excel
    2. Generate a targeted RAG question
    3. Retrieve relevant pages and extract the value via LLM
    4. Compare: if RAG value differs → mark as "changed"; if empty → mark as "filled"
    5. Generate a new Excel with color-coded changes

    Args:
        extraction_excel_path: Path to the extraction pipeline's output Excel.
        pdf_filename: If set, only use this PDF for RAG retrieval.
        statement_types: List of statement types to validate.
                        Default: ["balance_sheet", "profit_and_loss", "cash_flow"]
        session_id: Session ID for progress tracking.

    Returns:
        Dict with 'results', 'output_path', 'summary'.
    """
    if statement_types is None:
        statement_types = ["balance_sheet", "profit_and_loss", "cash_flow"]

    logger.info(
        f"CA RAG Agent: starting validation for '{extraction_excel_path}' "
        f"(pdf_filter={pdf_filename}, types={statement_types})"
    )

    # Initialize progress
    _progress_store[session_id] = {
        "status": "running",
        "current_step": "Reading extraction Excel",
        "progress": 0,
        "total_items": 0,
        "completed_items": 0,
        "details": [],
    }

    # Read original Excel values
    _update_progress(session_id, "Reading extraction Excel", 5, detail="Reading values from extraction Excel...")
    original_values = _read_extraction_excel(extraction_excel_path)
    _update_progress(session_id, "Reading extraction Excel", 10,
                     detail=f"Read {len(original_values)} values from Excel")

    # Read financial year from Meta Data sheet
    financial_year = _read_financial_year(extraction_excel_path)
    logger.info(f"CA RAG Agent: Financial Year from Meta Data: {financial_year}")
    _update_progress(session_id, "Reading financial year", 11,
                     detail=f"Financial Year: {financial_year or 'N/A'}")

    # Count total items for progress
    total_items = 0
    for stmt_type in statement_types:
        items_dict = _get_items_dict(stmt_type)
        if items_dict:
            total_items += sum(len(items) for items in items_dict.values())

    _update_progress(session_id, "Preparing validation", 12,
                     total=total_items, detail=f"Total {total_items} items to validate across {len(statement_types)} statements")

    # Run RAG validation for each statement type
    all_results: list[RAGValidationResult] = []
    completed_count = 0

    stmt_progress_labels = {
        "balance_sheet": "Validating Balance Sheet",
        "profit_and_loss": "Validating Profit & Loss",
        "cash_flow": "Validating Cash Flow",
    }

    for stmt_type in statement_types:
        items_dict = _get_items_dict(stmt_type)
        if not items_dict:
            continue

        questions = _generate_rag_questions(stmt_type, items_dict, financial_year=financial_year)
        step_label = stmt_progress_labels.get(stmt_type, f"Validating {stmt_type}")
        logger.info(f"CA RAG Agent: {stmt_type} — {len(questions)} items to validate")

        _update_progress(session_id, step_label, 12 + (completed_count / max(total_items, 1)) * 80,
                         detail=f"Starting {stmt_type}: {len(questions)} items")

        for q in questions:
            item = q["item"]
            section = q["section"]
            question = q["question"]

            # Get original value from extraction
            orig_val = original_values.get(item)

            # Skip formula/total items
            if _is_formula_item(item):
                all_results.append(RAGValidationResult(
                    item=item, section=section,
                    original_value=orig_val, rag_value=None,
                    action="skipped", confidence=0.0,
                    sources=[], raw_answer="Formula item — skipped",
                ))
                completed_count += 1
                continue

            # Ask RAG (with financial year context so LLM picks correct year column)
            rag_result = _extract_value_from_rag(
                question, filter_filename=pdf_filename, financial_year=financial_year
            )
            rag_val = rag_result["value"]

            # Determine action
            if rag_val is None:
                action = "not_found"
            elif orig_val is None or orig_val == "" or orig_val == "None":
                action = "filled"
            elif _values_match(orig_val, rag_val):
                action = "verified"
            else:
                action = "changed"

            all_results.append(RAGValidationResult(
                item=item, section=section,
                original_value=orig_val, rag_value=rag_val,
                action=action, confidence=rag_result["confidence"],
                sources=rag_result["sources"],
                raw_answer=rag_result["raw_answer"],
            ))

            completed_count += 1
            progress_pct = 12 + (completed_count / max(total_items, 1)) * 80

            # Update progress every 3 items or on interesting actions
            if action in ("changed", "filled") or completed_count % 3 == 0:
                _update_progress(session_id, step_label, progress_pct,
                                 completed=completed_count,
                                 detail=f"[{action}] {item}: orig={orig_val or '(empty)'}, rag={rag_val or 'N/A'}")

            logger.info(
                f"CA RAG: [{action}] {item}: "
                f"orig={orig_val}, rag={rag_val} "
                f"(confidence={rag_result['confidence']:.2f})"
            )

    # Generate color-coded Excel
    _update_progress(session_id, "Generating color-coded Excel", 95,
                     completed=completed_count,
                     detail=f"Writing Excel with {len(all_results)} validation results...")
    output_path = _generate_color_coded_excel(extraction_excel_path, all_results)

    # Summary
    summary = {
        "total_items": len(all_results),
        "verified": sum(1 for r in all_results if r.action == "verified"),
        "changed": sum(1 for r in all_results if r.action == "changed"),
        "filled": sum(1 for r in all_results if r.action == "filled"),
        "not_found": sum(1 for r in all_results if r.action == "not_found"),
        "skipped": sum(1 for r in all_results if r.action == "skipped"),
    }

    logger.info(f"CA RAG Agent: validation complete — {summary}")

    # Mark progress as complete
    _progress_store[session_id] = {
        "status": "completed",
        "current_step": "Validation complete",
        "progress": 100,
        "total_items": total_items,
        "completed_items": completed_count,
        "details": [f"Done: {summary['verified']} verified, {summary['changed']} changed, {summary['filled']} filled, {summary['not_found']} not found"],
    }

    return {
        "results": [
            {
                "item": r.item,
                "section": r.section,
                "original_value": r.original_value,
                "rag_value": r.rag_value,
                "action": r.action,
                "confidence": r.confidence,
                "sources": r.sources,
            }
            for r in all_results
        ],
        "output_path": output_path,
        "summary": summary,
    }


# ============================================================================
# EXCEL READING
# ============================================================================


def _read_extraction_excel(excel_path: str) -> dict[str, Optional[str]]:
    """
    Read values from the extraction pipeline's output Excel.

    Reads all sheets (BS, P&L, CF) and maps template item names to their values.

    The extraction Excel format is:
    - Column A: Section name (e.g., "Non-current assets")
    - Column B: Item name (e.g., "(a) Property, Plant and Equipment")
    - Column C+: Values (current year, previous year, etc.)

    Returns:
        Dict mapping template item name -> value string (or None if empty).
    """
    values = {}

    if not os.path.exists(excel_path):
        logger.warning(f"Extraction Excel not found: {excel_path}")
        return values

    # Known template item names for matching (to distinguish item names from section names)
    all_template_items = set()
    for items_dict in [BALANCE_SHEET_ITEMS, PL_ITEMS, CF_ITEMS]:
        for section, items in items_dict.items():
            for item in items:
                all_template_items.add(item)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            # Only read the main statement sheets (not Raw or Meta Data)
            if sheet_name in ("Raw BS", "Raw P&L", "Raw CF", "Meta Data", "RAG Legend", "RAG Details"):
                continue

            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, values_only=False):
                if not row or len(row) < 3:
                    continue

                # Try column B first (item name), then column A (for sheets where A has items)
                item_name = None
                value = None

                # Check column B (index 1) — this is the item name in the extraction format
                col_b = str(row[1].value).strip() if len(row) > 1 and row[1].value else ""
                # Check column A (index 0) — this is the section name, but could be item in some formats
                col_a = str(row[0].value).strip() if row[0].value else ""

                # Determine which column has the item name
                # We always want to store under the TEMPLATE item name (not the Excel name)
                # so that run_ca_rag_validation() can find values via original_values.get(item)
                template_key = None  # The template item name to use as key

                if col_b and col_b != "None" and col_b in all_template_items:
                    # Exact match in column B
                    template_key = col_b
                    # Value is in column C (index 2) or later
                    for cell in row[2:]:
                        if cell.value is not None:
                            cell_str = str(cell.value).strip()
                            if cell_str and cell_str != "None":
                                value = cell_str
                                break
                elif col_a and col_a != "None" and col_a in all_template_items:
                    # Exact match in column A
                    template_key = col_a
                    # Value is in column B (index 1) or later
                    for cell in row[1:]:
                        if cell.value is not None:
                            cell_str = str(cell.value).strip()
                            if cell_str and cell_str != "None":
                                value = cell_str
                                break
                elif col_b and col_b != "None":
                    # Column B has something that's not an exact template match
                    # Try fuzzy: check if any template item is a substring of col_b
                    # (Excel names may have extra text, e.g. "(iii) Other financial liabilities (other than...)")
                    for template_item in all_template_items:
                        if template_item in col_b:
                            template_key = template_item  # Store under TEMPLATE key
                            # Value is in column C (index 2) or later
                            for cell in row[2:]:
                                if cell.value is not None:
                                    cell_str = str(cell.value).strip()
                                    if cell_str and cell_str != "None":
                                        value = cell_str
                                        break
                            break
                    # Also try reverse: col_b is substring of template_item
                    if not template_key:
                        for template_item in all_template_items:
                            if col_b in template_item:
                                template_key = template_item
                                for cell in row[2:]:
                                    if cell.value is not None:
                                        cell_str = str(cell.value).strip()
                                        if cell_str and cell_str != "None":
                                            value = cell_str
                                            break
                                break

                if template_key:
                    values[template_key] = value

        wb.close()
    except Exception as e:
        logger.error(f"Failed to read extraction Excel: {e}")

    logger.info(f"Read {len(values)} values from extraction Excel")
    return values


def _read_financial_year(excel_path: str) -> Optional[str]:
    """
    Read the Financial Year from the Meta Data sheet of the extraction Excel.

    The Meta Data sheet has label-value pairs in columns A/B:
        Column A: Label (e.g., "Financial Year")
        Column B: Value (e.g., "2018" or "2017-18")

    Returns:
        Financial year string, or None if not found.
    """
    if not os.path.exists(excel_path):
        logger.warning(f"Extraction Excel not found: {excel_path}")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Meta Data" not in wb.sheetnames:
            logger.info("No 'Meta Data' sheet found in extraction Excel")
            wb.close()
            return None

        ws = wb["Meta Data"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and len(row) >= 2:
                label = str(row[0]).strip() if row[0] else ""
                if label == "Financial Year":
                    value = str(row[1]).strip() if row[1] else None
                    wb.close()
                    if value and value not in ("None", "N/A", ""):
                        logger.info(f"Read Financial Year from Meta Data: {value}")
                        return value
                    return None

        wb.close()
    except Exception as e:
        logger.warning(f"Could not read Financial Year from Meta Data: {e}")

    return None


# ============================================================================
# COLOR-CODED EXCEL GENERATION
# ============================================================================


def _generate_color_coded_excel(
    original_path: str,
    results: list[RAGValidationResult],
) -> str:
    """
    Generate a new Excel file based on the extraction pipeline output,
    with RAG-validated values and color coding applied IN-PLACE.

    The output keeps the SAME format as the extraction pipeline output
    (same sheets, same layout, same columns). The only changes are:
    1. Color coding applied to value cells based on RAG action
    2. RAG values written into cells that were changed or filled
    3. A "RAG Legend" worksheet added as an extra sheet

    Color scheme:
    - YELLOW:     RAG CHANGED a value (original had different value)
    - GREEN:      RAG FILLED a previously empty cell
    - LIGHT BLUE: RAG VERIFIED value matches (no change)
    - No color:   Not validated by RAG

    The output file is saved as: {original_name}_rag_validated.xlsx

    Returns:
        Path to the generated Excel file.
    """
    if not os.path.exists(original_path):
        logger.error(f"Original Excel not found: {original_path}")
        return ""

    # Build lookups: item -> RAGValidationResult
    result_lookup = {}
    for r in results:
        result_lookup[r.item] = r

    # Build lookup: item -> (rag_value, action) for changed/filled only
    rag_lookup = {}
    for r in results:
        if r.action in ("changed", "filled"):
            rag_lookup[r.item] = (r.rag_value, r.action)

    # Copy original workbook (preserves all sheets, formatting, formulas)
    wb = openpyxl.load_workbook(original_path)

    # ========================================================================
    # Apply RAG colors and values to existing sheets
    # ========================================================================
    # The extraction Excel format is:
    #   Column A: Section name (only on first row of each section, e.g., "Non-current assets")
    #   Column B: Item name (e.g., "(a) Property, Plant and Equipment")
    #   Column C: Value (current year)
    #   Column D+: Optional (previous year, etc.)
    #
    # We match items by Column B and write/colour values in Column C.
    # Fuzzy matching: Excel item names may have extra text vs template names
    # (e.g., "(iii) Other financial liabilities (other than those specified in item (b), to be specified)"
    #  vs template "(iii) Other financial liabilities")

    def _find_rag_match(excel_item_name: str, lookup: dict) -> Optional[str]:
        """Find a matching template item key for an Excel item name.
        
        Tries exact match first, then checks if any template key is a
        substring of the Excel name (to handle extra parenthetical text).
        """
        if not excel_item_name or excel_item_name == "None":
            return None
        # Exact match
        if excel_item_name in lookup:
            return excel_item_name
        # Substring match: template key is contained in Excel name
        for key in lookup:
            if key in excel_item_name:
                return key
        # Substring match: Excel name is contained in template key
        for key in lookup:
            if excel_item_name in key:
                return key
        return None

    sheets_to_process = [s for s in wb.sheetnames
                         if s not in ("Meta Data", "Raw BS", "Raw P&L", "Raw CF", "RAG Legend", "RAG Details")]

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        for row_num in range(1, ws.max_row + 1):
            # Read column B (item name)
            col_b_val = ws.cell(row=row_num, column=2).value
            excel_item = str(col_b_val).strip() if col_b_val else ""

            if not excel_item or excel_item == "None":
                continue

            # Find matching RAG result key (exact or fuzzy)
            rag_key = _find_rag_match(excel_item, rag_lookup)
            result_key = _find_rag_match(excel_item, result_lookup)

            # Check if this item has a RAG result (changed or filled)
            if rag_key and rag_key in rag_lookup:
                rag_value, action = rag_lookup[rag_key]

                # Write RAG value to column C
                value_cell = ws.cell(row=row_num, column=3)
                value_cell.value = rag_value

                # Apply color
                if action == "changed":
                    value_cell.fill = FILL_CHANGED
                    value_cell.font = Font(bold=True, color="CC0000")  # Red bold for changed
                elif action == "filled":
                    value_cell.fill = FILL_NEW
                    value_cell.font = Font(bold=True, color="006100")  # Dark green bold for filled

            # Check for verified items (apply light blue to value cell in column C)
            if result_key and result_key in result_lookup:
                r = result_lookup[result_key]
                if r.action == "verified":
                    value_cell = ws.cell(row=row_num, column=3)
                    if value_cell.value is not None:
                        value_cell.fill = FILL_VERIFIED

    # ========================================================================
    # Add RAG Legend sheet (extra worksheet)
    # ========================================================================
    legend_ws = wb.create_sheet("RAG Legend", 0)
    legend_ws["A1"] = "CA RAG Agent — Validation Report"
    legend_ws["A1"].font = Font(bold=True, size=16, color="1F4E79")

    # Color Legend
    legend_ws["A3"] = "Color Legend:"
    legend_ws["A3"].font = Font(bold=True, size=12)
    legend_ws["A4"].fill = FILL_CHANGED
    legend_ws["B4"] = "YELLOW — RAG CHANGED a value (original was different)"
    legend_ws["A5"].fill = FILL_NEW
    legend_ws["B5"] = "GREEN — RAG FILLED an empty cell (new data from RAG)"
    legend_ws["A6"].fill = FILL_VERIFIED
    legend_ws["B6"] = "LIGHT BLUE — RAG VERIFIED value matches (no change needed)"

    # Summary statistics
    changed = sum(1 for r in results if r.action == "changed")
    filled = sum(1 for r in results if r.action == "filled")
    verified = sum(1 for r in results if r.action == "verified")
    not_found = sum(1 for r in results if r.action == "not_found")
    skipped = sum(1 for r in results if r.action == "skipped")

    legend_ws["A8"] = "Validation Summary:"
    legend_ws["A8"].font = Font(bold=True, size=12)
    legend_ws["A9"] = f"Values changed by RAG: {changed}"
    legend_ws["A10"] = f"Empty cells filled by RAG: {filled}"
    legend_ws["A11"] = f"Values verified (match): {verified}"
    legend_ws["A12"] = f"Values not found by RAG: {not_found}"
    legend_ws["A13"] = f"Skipped (formula items): {skipped}"

    # Highlight changed/filled counts
    if changed > 0:
        legend_ws["A9"].font = Font(bold=True, color="CC0000")
    if filled > 0:
        legend_ws["A10"].font = Font(bold=True, color="006100")

    # Key insight
    legend_ws["A15"] = "Key Insight:"
    legend_ws["A15"].font = Font(bold=True, size=12)
    legend_ws["A16"] = (
        f"RAG improved the extraction: {changed} values corrected, "
        f"{filled} empty cells filled, {verified} values confirmed."
    )

    # Per-statement breakdown
    stmt_labels = {
        "balance_sheet": "Balance Sheet",
        "profit_and_loss": "Profit & Loss",
        "cash_flow": "Cash Flow",
    }
    legend_ws["A18"] = "Per-Statement Breakdown:"
    legend_ws["A18"].font = Font(bold=True, size=12)

    # Group results by statement type
    stmt_results = {}
    for r in results:
        # Determine statement type from section
        stmt_type = None
        for st, items_dict in [("balance_sheet", BALANCE_SHEET_ITEMS),
                                ("profit_and_loss", PL_ITEMS),
                                ("cash_flow", CF_ITEMS)]:
            for section, items in items_dict.items():
                if r.item in items:
                    stmt_type = st
                    break
            if stmt_type:
                break

        if stmt_type:
            if stmt_type not in stmt_results:
                stmt_results[stmt_type] = []
            stmt_results[stmt_type].append(r)

    row_offset = 19
    for stmt_type, label in stmt_labels.items():
        stmt_res = stmt_results.get(stmt_type, [])
        v = sum(1 for r in stmt_res if r.action == "verified")
        c = sum(1 for r in stmt_res if r.action == "changed")
        f = sum(1 for r in stmt_res if r.action == "filled")
        nf = sum(1 for r in stmt_res if r.action == "not_found")
        sk = sum(1 for r in stmt_res if r.action == "skipped")
        legend_ws.cell(row=row_offset, column=1,
            value=f"  {label}: {v} verified, {c} changed, {f} filled, {nf} not found, {sk} skipped")
        row_offset += 1

    # Value source explanation
    legend_ws.cell(row=row_offset + 1, column=1, value="How a CA determines each value:").font = Font(bold=True, size=12)
    FILL_DIRECT = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    FILL_CALCULATED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    FILL_INFERRED = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    legend_ws.cell(row=row_offset + 2, column=1).fill = FILL_DIRECT
    legend_ws.cell(row=row_offset + 2, column=2, value="DIRECT — Value explicitly stated in the financial statement")
    legend_ws.cell(row=row_offset + 3, column=1).fill = FILL_CALCULATED
    legend_ws.cell(row=row_offset + 3, column=2, value="CALCULATED — Sum of sub-items (e.g., Financial Assets = Investments + Trade Receivables + Loans + Others)")
    legend_ws.cell(row=row_offset + 4, column=1).fill = FILL_INFERRED
    legend_ws.cell(row=row_offset + 4, column=2, value="INFERRED — Derived from accounting equations (e.g., BS: Assets = Equity + Liabilities)")

    legend_ws.column_dimensions['A'].width = 55
    legend_ws.column_dimensions['B'].width = 80

    # ========================================================================
    # Add RAG Details sheet with all validation results
    # ========================================================================
    details_ws = wb.create_sheet("RAG Details")
    headers = ["Item", "Section", "Original Value", "RAG Value", "Action",
               "Confidence", "Source Pages"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = details_ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    for i, r in enumerate(results, 2):
        details_ws.cell(row=i, column=1, value=r.item)
        details_ws.cell(row=i, column=2, value=r.section)
        details_ws.cell(row=i, column=3, value=r.original_value or "")
        details_ws.cell(row=i, column=4, value=r.rag_value or "")
        details_ws.cell(row=i, column=5, value=r.action)
        details_ws.cell(row=i, column=6, value=round(r.confidence, 2) if r.confidence else "")
        source_pages = ", ".join(
            f"p{s['page']}({s['filename']})" for s in r.sources
        )
        details_ws.cell(row=i, column=7, value=source_pages)

        # Color the action cell
        action_cell = details_ws.cell(row=i, column=5)
        if r.action == "changed":
            action_cell.fill = FILL_CHANGED
        elif r.action == "filled":
            action_cell.fill = FILL_NEW
        elif r.action == "verified":
            action_cell.fill = FILL_VERIFIED

    details_ws.column_dimensions['A'].width = 50
    details_ws.column_dimensions['B'].width = 25
    details_ws.column_dimensions['C'].width = 20
    details_ws.column_dimensions['D'].width = 20
    details_ws.column_dimensions['E'].width = 14
    details_ws.column_dimensions['F'].width = 12
    details_ws.column_dimensions['G'].width = 35

    # Save
    base, ext = os.path.splitext(original_path)
    output_path = f"{base}_rag_validated{ext}"
    wb.save(output_path)
    wb.close()

    logger.info(f"RAG-validated Excel saved (same format + RAG Legend + RAG Details): {output_path}")
    return output_path


# ============================================================================
# HELPERS
# ============================================================================


def _get_items_dict(statement_type: str) -> dict:
    """Get the template items dict for a statement type."""
    return {
        "balance_sheet": BALANCE_SHEET_ITEMS,
        "profit_and_loss": PL_ITEMS,
        "cash_flow": CF_ITEMS,
    }.get(statement_type, {})


def generate_prompt_excel(output_path: str = None) -> str:
    """
    Generate a reference Excel showing exactly what questions the CA RAG agent
    will ask for each template item. This lets users review the prompts before
    running validation.

    The Excel has one sheet per statement type (BS, P&L, CF) with columns:
        - Section
        - Template Item
        - Value Source (Direct / Calculated / Inferred)
        - RAG Question (exact prompt sent to LLM)
        - Purpose (verify / fill)

    Args:
        output_path: Where to save. Default: api_output/rag_prompt_reference.xlsx

    Returns:
        Path to the generated Excel file.
    """
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "api_output", "rag_prompt_reference.xlsx"
        )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()

    sheet_configs = [
        ("Balance Sheet", "balance_sheet", BALANCE_SHEET_ITEMS),
        ("Profit & Loss", "profit_and_loss", PL_ITEMS),
        ("Cash Flow", "cash_flow", CF_ITEMS),
    ]

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Color fills for value source classification
    FILL_DIRECT = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")      # Light blue
    FILL_CALCULATED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Light yellow
    FILL_INFERRED = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Light green

    for sheet_name, stmt_type, items_dict in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        headers = ["#", "Section", "Template Item", "Value Source", "RAG Question (exact prompt)", "Purpose"]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        # Generate questions (without year context — year is dynamic per PDF,
        # added at runtime from the extraction Excel's Meta Data sheet)
        questions = _generate_rag_questions(stmt_type, items_dict, financial_year=None)

        for i, q in enumerate(questions, 2):
            item = q["item"]
            section = q["section"]
            question = q["question"]
            value_source = q.get("value_source", "Direct")

            # Determine purpose
            if _is_formula_item(item):
                purpose = "Skipped (formula item)"
            else:
                purpose = "Verify if filled, Fill if empty"

            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=section)
            ws.cell(row=i, column=3, value=item)

            # Value source cell with color coding
            vs_cell = ws.cell(row=i, column=4, value=value_source)
            if "Direct" in value_source and "Inferred" not in value_source:
                vs_cell.fill = FILL_DIRECT
            elif "Calculated" in value_source:
                vs_cell.fill = FILL_CALCULATED
            elif "Inferred" in value_source:
                vs_cell.fill = FILL_INFERRED

            ws.cell(row=i, column=5, value=question)
            ws.cell(row=i, column=6, value=purpose)

            # Alternate row coloring (light)
            if i % 2 == 0:
                light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col in [1, 2, 3, 5, 6]:  # Skip value source column (has its own color)
                    ws.cell(row=i, column=col).fill = light_fill

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 100
        ws.column_dimensions['F'].width = 25

    # Add Summary sheet
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws["A1"] = "CA RAG Agent — Prompt Reference"
    summary_ws["A1"].font = Font(bold=True, size=14)

    summary_ws["A3"] = "What this Excel shows:"
    summary_ws["A3"].font = Font(bold=True, size=11)
    summary_ws["A4"] = "Each sheet lists the EXACT questions the CA RAG Agent will ask when validating extraction results."
    summary_ws["A5"] = "The agent does TWO things for each template item:"
    summary_ws["A6"] = "  1. VERIFY: If the extraction pipeline already filled a value, RAG cross-checks it against the PDF"
    summary_ws["A7"] = "  2. FILL: If the extraction pipeline left a cell empty, RAG tries to find the value from the PDF"

    summary_ws["A9"] = "How a CA determines each value (Value Source column):"
    summary_ws["A9"].font = Font(bold=True, size=11)
    summary_ws["A10"].fill = FILL_DIRECT
    summary_ws["B10"] = "DIRECT — Value is explicitly stated in the financial statement (e.g., 'Property, Plant & Equipment: 6,527.97')"
    summary_ws["A11"].fill = FILL_CALCULATED
    summary_ws["B11"] = "CALCULATED — Value is a sum/subtotal of sub-items (e.g., 'Financial Assets' = Investments + Trade Receivables + Loans + Others)"
    summary_ws["A12"].fill = FILL_INFERRED
    summary_ws["B12"] = "INFERRED — Value derived from accounting equations (e.g., BS: Assets = Equity + Liabilities; P&L: Profit = Income - Expenses)"
    summary_ws["A13"] = "The RAG agent thinks like a CA: it looks for direct values first, then calculates from sub-items, then infers from equations."

    summary_ws["A15"] = "Color coding in RAG-validated output:"
    summary_ws["A15"].font = Font(bold=True, size=11)
    summary_ws["A16"].fill = FILL_CHANGED
    summary_ws["B16"] = "YELLOW — RAG CHANGED a value (original was different)"
    summary_ws["A17"].fill = FILL_NEW
    summary_ws["B17"] = "GREEN — RAG FILLED an empty cell (new data)"
    summary_ws["A18"].fill = FILL_VERIFIED
    summary_ws["B18"] = "LIGHT BLUE — RAG VERIFIED value matches (no change)"

    summary_ws["A20"] = "Statement Type Detection:"
    summary_ws["A20"].font = Font(bold=True, size=11)
    summary_ws["A21"] = "Statement type (Standalone vs Consolidated) is AUTO-DETECTED by the extraction pipeline"
    summary_ws["A22"] = "using a 3-layer cascade: LLM primary → deterministic regex → cross-validation."
    summary_ws["A23"] = "The RAG agent only validates STANDALONE financial statements (per policy)."

    summary_ws["A25"] = "Financial Year Context:"
    summary_ws["A25"].font = Font(bold=True, size=11)
    summary_ws["A26"] = "The financial year is DYNAMIC — read from the extraction Excel's Meta Data sheet at runtime."
    summary_ws["A27"] = "Questions shown above do NOT include year context. At runtime, each question is augmented with:"
    summary_ws["A28"] = '  - Year context: e.g., "for FY 2017-18"'
    summary_ws["A29"] = '  - Column instruction: e.g., "Extract the value from the column for the year ending March 31, 2018 (current year). Do NOT use the previous year column."'
    summary_ws["A30"] = "This ensures the RAG agent extracts values for the CORRECT financial year only."

    summary_ws["A32"] = "Total items per statement:"
    summary_ws["A32"].font = Font(bold=True, size=11)
    for i, (sheet_name, stmt_type, items_dict) in enumerate(sheet_configs):
        total = sum(len(items) for items in items_dict.values())
        direct = len(VALUE_SOURCE_CLASSIFICATION.get(stmt_type, {}).get("direct", []))
        calculated = len(VALUE_SOURCE_CLASSIFICATION.get(stmt_type, {}).get("calculated", []))
        inferred = len(VALUE_SOURCE_CLASSIFICATION.get(stmt_type, {}).get("inferred", []))
        summary_ws.cell(row=33 + i, column=1,
            value=f"  {sheet_name}: {total} items ({direct} direct, {calculated} calculated, {inferred} inferred)")

    summary_ws.column_dimensions['A'].width = 55
    summary_ws.column_dimensions['B'].width = 80

    wb.save(output_path)
    wb.close()

    logger.info(f"RAG Prompt Reference Excel saved: {output_path}")
    return output_path


def _is_formula_item(item: str) -> bool:
    """Check if an item is a formula/total row that should be skipped."""
    formula_items = {
        "Total Non Current Assets", "Total Fixed assets", "Total Current Assets",
        "Total Assets", "Total Equity", "Total Non-current Liabilities",
        "Total Current Liabilities", "Total Liabilities",
        "Total Equity and Liabilities", "Total debt",
        "(h) Financial Assets", "(b) Financial Assets",
        "(a) Financial Liabilities", "(ii) Trade Payables",
        "III. Total Income (I + II)", "VIII. Tax expense",
        "Total expenses", "V. Profit/(loss) before exceptional items and tax (III - IV)",
        "VII. Profit/(loss) before tax (V-VI)",
        "IX. Profit/(Loss) for the period from continuing operations (VII-VIII)",
        "XIII. Profit/(Loss) after taxes (IX + XII)",
        "EBITDA", "EBIT",
    }
    return item in formula_items


def _values_match(orig: str, rag: str) -> bool:
    """
    Check if original and RAG values match (after normalization).

    Handles:
    - Comma differences: "6,527.97" vs "6527.97"
    - Parentheses vs minus: "(1,234.56)" vs "-1234.56"
    - Whitespace differences
    """
    if orig is None or rag is None:
        return False

    def normalize(s):
        s = str(s).strip()
        s = s.replace(",", "").replace(" ", "")
        # Normalize parentheses to minus
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        return s

    return normalize(orig) == normalize(rag)
