from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .config import AppConfig, get_config

DEFAULT_EXCEL_LABEL_ALIASES = {
    "Property Plant and Equipment": ["Property Plant and Equipment", "PPE", "Property, Plant & Equipment", "Fixed Assets"],
    "Property Plant Equipment": ["Property Plant Equipment", "PPE", "Property, Plant & Equipment", "Fixed Assets"],
    "Total Assets": ["Total Assets", "Total assets"],
    "Total Liabilities": ["Total Liabilities", "Total liabilities"],
    "Total Equity": ["Total Equity", "Total equity"],
    "Profit before tax": ["Profit before tax", "Profit Before Tax"],
    "Profit after tax": ["Profit after tax", "Profit After Tax"],
    "Operating Cash Flow": ["Operating Cash Flow", "Cash from Operations"],
    "Investing Cash Flow": ["Investing Cash Flow", "Cash from Investing"],
    "Financing Cash Flow": ["Financing Cash Flow", "Cash from Financing"],
}

logger = logging.getLogger(__name__)


def _normalize_label(label: str) -> str:
    """Normalize labels for alias matching."""
    return re.sub(r"\s+", " ", str(label).strip().lower())


def _get_excel_label_aliases(label: str, custom_aliases: Optional[Dict[str, list[str]]] = None) -> set[str]:
    """Return a normalized set of label aliases for matching."""
    aliases = set()
    alias_source = custom_aliases or DEFAULT_EXCEL_LABEL_ALIASES
    for key, values in alias_source.items():
        if _normalize_label(label) in {_normalize_label(value) for value in values}:
            aliases.add(_normalize_label(key))
    aliases.add(_normalize_label(label))
    return aliases


def resolve_excel_target_cell(worksheet: Any, label: str, custom_aliases: Optional[Dict[str, list[str]]] = None) -> Any:
    """Find the first matching cell in the worksheet for a label or alias without hardcoded row numbers."""
    normalized_label = _normalize_label(label)
    aliases = _get_excel_label_aliases(label, custom_aliases)

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if _normalize_label(str(cell.value)) in aliases:
                return cell

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if normalized_label in _normalize_label(str(cell.value)) or _normalize_label(str(cell.value)) in aliases:
                return cell

    raise KeyError(f"No matching Excel label found for: {label}")


def write_excel_output(
    validated_entities_path: Optional[Path | str] = None,
    template_path: Optional[Path | str] = None,
    output_path: Optional[Path | str] = None,
    config: Optional[AppConfig] = None,
) -> Path:
    """Populate an Excel template with validated financial entities.

    The source template workbook is opened read-only and copied into a new workbook
    so the original file remains unchanged. Matching labels in any worksheet are
    populated into the adjacent cell.
    """
    app_config = config or get_config()

    input_path = Path(validated_entities_path) if validated_entities_path else app_config.output_folder / "validated_entities.json"
    template_file = Path(template_path) if template_path else app_config.template_folder / "financial_template.xlsx"
    target_path = Path(output_path) if output_path else app_config.output_folder / "financial_output.xlsx"

    if not input_path.exists():
        raise FileNotFoundError(f"Validated entities file not found: {input_path}")
    if not template_file.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_file}")

    with input_path.open("r", encoding="utf-8") as handle:
        entities: Dict[str, Any] = json.load(handle)

    workbook = load_workbook(filename=str(template_file), data_only=False)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                label = str(cell.value).strip()
                matched_key = None
                for entity_key in entities:
                    if entity_key is None:
                        continue
                    if _normalize_label(entity_key) == _normalize_label(label):
                        matched_key = entity_key
                        break
                    if _normalize_label(entity_key) in _get_excel_label_aliases(label):
                        matched_key = entity_key
                        break

                if matched_key is None:
                    continue

                value = entities[matched_key]
                if value is None:
                    continue
                if isinstance(value, str) and value.startswith("{"):
                    continue

                target_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                target_cell.value = value

    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    logger.info("Excel output written to %s", target_path)
    return target_path
