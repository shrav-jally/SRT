from pathlib import Path
from openpyxl import Workbook

input_dir = Path('input')
input_dir.mkdir(exist_ok=True)
pdf_path = input_dir / 'sample_annual_report.pdf'
pdf_path.write_bytes(b'%PDF-1.4\n1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 300 144] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >> endobj\n4 0 obj << /Length 44 >> stream\nBT /F1 18 Tf 50 100 Td (Sample Annual Report) Tj ET\nendstream\nendobj\n5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\nxref\n0 6\n0000000000 65535 f \n0000000010 00000 n \n0000000062 00000 n \n0000000119 00000 n \n0000000206 00000 n \n0000000305 00000 n \ntrailer << /Size 6 /Root 1 0 R >>\nstartxref\n0\n%%EOF\n')

wb = Workbook()
ws = wb.active
ws.title = 'Financials'
ws['A1'] = 'Company Name'
ws['B1'] = ''
ws['A2'] = 'Revenue from operations'
ws['B2'] = ''
ws['A3'] = 'EBITDA'
ws['B3'] = ''
wb.save('templates/financial_template.xlsx')
print('created sample files')
