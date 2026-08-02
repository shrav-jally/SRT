"""Reconstruct financial tables into Excel workbooks (Category-based sheets).

Sprint 2 additions:
  - Valuation Counterpart sheet (47-parameter extraction matrix)
  - Enhanced Intelligence Report with traceability columns
  - Colorful conditional formatting and group-based color coding
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore[import-not-found]
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ===================================================================
# Style constants
# ===================================================================

_FONT_TITLE = Font(name="Calibri", size=16, bold=True)
_FONT_SUBTITLE = Font(name="Calibri", size=14, bold=True, italic=True)
_FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="1F4E79")
_FONT_SUBSECTION = Font(name="Calibri", size=10, bold=True, color="2E75B6")
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_DATA = Font(name="Calibri", size=10)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True)

_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_SECTION = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_FILL_TOTAL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_ALT_ROW = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

_BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ===================================================================
# Sprint 2: Intelligence Report color palette
# ===================================================================

# Category group fills (for Intelligence Report)
_FILL_COMPANY_INFO = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")      # Light teal
_FILL_MGMT_GOV = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")           # Light purple
_FILL_SHAREHOLDING = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")       # Light orange
_FILL_MDA = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")                # Light green

# Status fills
_FILL_FOUND = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")              # Green
_FILL_NOT_DISCLOSED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Red
_FILL_NOT_APPLICABLE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # Yellow

# Status fonts
_FONT_FOUND = Font(name="Calibri", size=10, bold=True, color="006100")
_FONT_NOT_DISCLOSED = Font(name="Calibri", size=10, bold=True, color="9C0006")
_FONT_NOT_APPLICABLE = Font(name="Calibri", size=10, bold=True, color="9C6500")

# Traceability column fill
_FILL_TRACE = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_FONT_TRACE_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FILL_TRACE_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")  # Olive green

# ===================================================================
# Sprint 2: Valuation Counterpart color palette
# ===================================================================

# Group fills for Valuation Counterpart
_FILL_VC_METADATA = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")         # Blue
_FILL_VC_PNL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")              # Orange
_FILL_VC_BS = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")               # Green
_FILL_VC_SHARES = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")           # Gold
_FILL_VC_CLASS = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")            # Light blue
_FILL_VC_OPTIONAL = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")         # Gray

# Light versions for row backgrounds
_FILL_VC_METADATA_LT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_FILL_VC_PNL_LT = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_FILL_VC_BS_LT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_VC_SHARES_LT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_VC_CLASS_LT = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_FILL_VC_OPTIONAL_LT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Valuation header
_FILL_VC_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_FONT_VC_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

# Valuation traceability header
_FILL_VC_TRACE_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

# Requirement fills
_FILL_REQ = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Green for Req
_FILL_OPT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")     # Light blue for Opt

# ===================================================================
# Category -> fill mapping for Intelligence Report
# ===================================================================

_INTEL_CATEGORY_FILLS = {
    "Company Information": _FILL_COMPANY_INFO,
    "Management & Governance": _FILL_MGMT_GOV,
    "Shareholding Information": _FILL_SHAREHOLDING,
    "Management Discussion & Analysis": _FILL_MDA,
}

# ===================================================================
# Group -> fill mapping for Valuation Counterpart
# ===================================================================

_VC_GROUP_HEADER_FILLS = {
    "Metadata": _FILL_VC_METADATA,
    "P&L": _FILL_VC_PNL,
    "Balance Sheet": _FILL_VC_BS,
    "Shares": _FILL_VC_SHARES,
    "Classification": _FILL_VC_CLASS,
    "Optional": _FILL_VC_OPTIONAL,
}

_VC_GROUP_ROW_FILLS = {
    "Metadata": _FILL_VC_METADATA_LT,
    "P&L": _FILL_VC_PNL_LT,
    "Balance Sheet": _FILL_VC_BS_LT,
    "Shares": _FILL_VC_SHARES_LT,
    "Classification": _FILL_VC_CLASS_LT,
    "Optional": _FILL_VC_OPTIONAL_LT,
}


# ===================================================================
# Main builder
# ===================================================================


PRIORITY_MAP = {
    # Priority 1 (Ignored/Hold)
    "Financial Statements": 1,
    "Notes to Accounts": 1,
    
    # Priority 2
    "Company Information": 2,
    "Management & Governance": 2,
    "Shareholding Information": 2,
    "Management Discussion & Analysis": 2,
    
    # Priority 3
    "Investor Information": 3,
    "Audit Information": 3,
    "Outlook & Guidance": 3,
    
    # Priority 4
    "Financial Analysis": 4,
    "Business Performance": 4,
    "Risk Management": 4,
    
    # Priority 5
    "Legal & Compliance": 5,
    "Strategic Initiatives": 5,
    
    # Priority 6
    "ESG & Sustainability": 6,
    "CSR": 6,
    
    # Priority 7
    "Human Resources": 7,
}

def build_excel(extraction_result: dict[str, Any]) -> bytes:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    metadata = extraction_result.get("metadata", {})
    master_sections = extraction_result.get("master_sections", [])
    text_extractions = extraction_result.get("text_extractions", [])
    table_extractions = extraction_result.get("table_extractions", [])

    _build_metadata_sheet(wb, extraction_result)
    
    if master_sections:
        _build_master_sections_sheet(wb, master_sections)
    if extraction_result.get("table_inventory"):
        _build_table_inventory_sheet(wb, extraction_result["table_inventory"])

    # Build Category Sheets
    categories = {}
    for sec in master_sections:
        cat = sec.get("category", "Uncategorized")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(sec)

    for cat, sections in categories.items():
        _build_category_sheet(wb, cat, sections, text_extractions, table_extractions)

    # Build Priority sheets (Sprint 2 / Intelligence Report equivalent)
    structured_intel = extraction_result.get("structured_intelligence", {})
    evidence_map = extraction_result.get("evidence_map", {})
    if structured_intel:
        try:
            from .workbook_population import populate_intelligence_report
            priority_buckets = {}
            for cat, subcategories in structured_intel.items():
                priority = PRIORITY_MAP.get(cat, 2)  # Default unmapped to 2
                if priority == 1:
                    continue
                if priority not in priority_buckets:
                    priority_buckets[priority] = {}
                priority_buckets[priority][cat] = subcategories

            for priority in sorted(priority_buckets.keys()):
                p_intel = priority_buckets[priority]
                report_rows = populate_intelligence_report(p_intel, master_sections, evidence_map)
                _build_intelligence_sheet(wb, report_rows, priority)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning(f"Failed to build priority sheets: {exc}")

    # Build Valuation Counterpart sheet (Sprint 2)
    if table_extractions or structured_intel:
        try:
            from .workbook_population import populate_valuation_report
            valuation_rows = populate_valuation_report(
                table_extractions=table_extractions,
                structured_intelligence=structured_intel,
                master_sections=master_sections,
                metadata=metadata,
                raw_pages_text=extraction_result.get("raw_pages_text"),
            )
            _build_valuation_counterpart_sheet(wb, valuation_rows)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning(f"Failed to build valuation counterpart sheet: {exc}")

    # Move sheets to desired order:
    priority_sheets = ["Metadata", "Priority 2", "Valuation Counterpart", "Priority 3", "Priority 4", "Priority 5", "Priority 6", "Priority 7"]
    for i, sheet_name in enumerate(priority_sheets):
        if sheet_name in wb.sheetnames:
            current_idx = wb.sheetnames.index(sheet_name)
            if current_idx != i:
                wb.move_sheet(sheet_name, offset=i - current_idx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(extraction_result: dict[str, Any], path: str | Path) -> Path:
    path = Path(path)
    path.write_bytes(build_excel(extraction_result))
    return path


# ===================================================================
# Intelligence Sheet Builder (Sprint 2 Enhanced)
# ===================================================================

def _build_intelligence_sheet(wb: "openpyxl.Workbook", rows: list[dict], priority: int = 2) -> None:
    ws = wb.create_sheet(title=f"Priority {priority}")
    
    # == Title row ==
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    title_cell = ws.cell(row=1, column=1, value=f"PRIORITY {priority} -- Narrative Extraction Matrix")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # == Headers ==
    headers = [
        ("Category", 22),
        ("Sub Category", 28),
        ("Status", 16),
        ("Extracted Value", 70),
        ("Source Page", 14),
        ("Confidence", 12),
        ("Extraction Method", 30),
        ("Valuation Link", 25),
        ("Source Derivation", 45),
        ("Evidence Method", 20),
        ("Evidence Snippet", 50),
        ("Evidence Page", 14),
    ]
    
    header_row = 2
    for col_idx, (h, _) in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
    
    # Traceability sub-headers (columns 7-9 get a different header color)
    for col_idx in range(7, 10):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = _FONT_TRACE_HEADER
        c.fill = _FILL_TRACE_HEADER
    
    # Evidence sub-headers (columns 10-12 get a distinct color)
    _FILL_EVIDENCE_HEADER = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    _FONT_EVIDENCE_HEADER = Font(name="Calibri", size=10, bold=True, color="375623")
    for col_idx in range(10, 13):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = _FONT_EVIDENCE_HEADER
        c.fill = _FILL_EVIDENCE_HEADER
    
    # Evidence data fill
    _FILL_EVIDENCE = PatternFill(start_color="F2F7ED", end_color="F2F7ED", fill_type="solid")
    
    # == Data rows ==
    for row_idx, r in enumerate(rows, 3):
        cat = r["category"]
        subcat = r["subcategory"]
        status = r.get("status", "")
        
        # Category cell with group color
        cat_cell = ws.cell(row=row_idx, column=1, value=cat)
        cat_cell.alignment = _ALIGN_TOP_LEFT
        cat_fill = _INTEL_CATEGORY_FILLS.get(cat)
        if cat_fill:
            cat_cell.fill = cat_fill
        cat_cell.border = _BORDER_THIN
        
        # Subcategory
        sub_cell = ws.cell(row=row_idx, column=2, value=subcat)
        sub_cell.alignment = _ALIGN_TOP_LEFT
        if cat_fill:
            sub_cell.fill = cat_fill
        sub_cell.border = _BORDER_THIN
        
        # Status with conditional color
        status_cell = ws.cell(row=row_idx, column=3, value=status)
        status_cell.alignment = _ALIGN_CENTER
        status_cell.border = _BORDER_THIN
        if status == "FOUND":
            status_cell.fill = _FILL_FOUND
            status_cell.font = _FONT_FOUND
        elif status == "NOT DISCLOSED":
            status_cell.fill = _FILL_NOT_DISCLOSED
            status_cell.font = _FONT_NOT_DISCLOSED
        elif status == "NOT APPLICABLE":
            status_cell.fill = _FILL_NOT_APPLICABLE
            status_cell.font = _FONT_NOT_APPLICABLE
        
        # Extracted value
        val_cell = ws.cell(row=row_idx, column=4, value=r["extracted_value"])
        val_cell.alignment = _ALIGN_LEFT
        val_cell.border = _BORDER_THIN
        
        # Source page
        ws.cell(row=row_idx, column=5, value=r["source_page"]).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=5).border = _BORDER_THIN
        
        # Confidence
        ws.cell(row=row_idx, column=6, value=r["confidence"]).alignment = _ALIGN_CENTER
        ws.cell(row=row_idx, column=6).border = _BORDER_THIN
        
        # Traceability columns (7-9)
        ws.cell(row=row_idx, column=7, value=r.get("extraction_method", "")).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=7).border = _BORDER_THIN
        ws.cell(row=row_idx, column=7).fill = _FILL_TRACE
        
        ws.cell(row=row_idx, column=8, value=r.get("valuation_link", "")).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=8).border = _BORDER_THIN
        ws.cell(row=row_idx, column=8).fill = _FILL_TRACE
        
        ws.cell(row=row_idx, column=9, value=r.get("source_derivation", "")).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=9).border = _BORDER_THIN
        ws.cell(row=row_idx, column=9).fill = _FILL_TRACE
        
        # Evidence columns (10-12) — GAP 0C
        ws.cell(row=row_idx, column=10, value=r.get("evidence_method", "")).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=10).border = _BORDER_THIN
        ws.cell(row=row_idx, column=10).fill = _FILL_EVIDENCE
        
        ws.cell(row=row_idx, column=11, value=r.get("evidence_snippet", "")).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=11).border = _BORDER_THIN
        ws.cell(row=row_idx, column=11).fill = _FILL_EVIDENCE
        
        ev_page = r.get("evidence_page")
        ws.cell(row=row_idx, column=12, value=f"Page {ev_page}" if ev_page else "").alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=12).border = _BORDER_THIN
        ws.cell(row=row_idx, column=12).fill = _FILL_EVIDENCE
        
        # Auto-height for long values
        val_len = len(str(r["extracted_value"]))
        if val_len > 100:
            ws.row_dimensions[row_idx].height = max(30, min(200, val_len // 3))
    
    # == Column widths ==
    col_widths = [22, 28, 16, 70, 14, 12, 30, 25, 45, 20, 50, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze panes
    ws.freeze_panes = "A3"
    
    # Auto-filter
    ws.auto_filter.ref = f"A2:L{len(rows) + 2}"


# ===================================================================
# Valuation Counterpart Sheet Builder (Sprint 2)
# ===================================================================

def _build_valuation_counterpart_sheet(wb: "openpyxl.Workbook", rows: list[dict]) -> None:
    ws = wb.create_sheet(title="Valuation Counterpart")
    
    # == Title row ==
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    title_cell = ws.cell(row=1, column=1, value="VALUATION COUNTERPART -- 47-Parameter Extraction Matrix")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # == Subtitle / legend row ==
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    legend_cell = ws.cell(row=2, column=1,
        value="RAW reported line items from filed statements (Ind AS / Schedule III). Engine derives all ratios. "
              "Green = FOUND | Red = NOT DISCLOSED | Yellow = NOT APPLICABLE")
    legend_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25
    
    # == Headers ==
    # Columns: #, Group, Field, Statement/Section, Req?, Both Yrs, Type, Engine Use,
    #          Current Value, Previous Value, Status,
    #          Source Statement, Source Page, Note No, Raw Line Item,
    #          Trace: Source, Trace: Method, Trace: Intel Link, Trace: Derivation
    headers = [
        ("#", 5),
        ("Group", 14),
        ("Field (canonical label)", 38),
        ("Statement / Section", 28),
        ("Req?", 6),
        ("Both Yrs", 8),
        ("Type", 8),
        ("Feeds (engine use)", 30),
        ("Current Year Value", 18),
        ("Previous Year Value", 18),
        ("YoY Growth %", 14),
        ("Status", 16),
        ("Source Statement", 22),
        ("Source Page", 12),
        ("Note No", 8),
        ("Raw Line Item", 25),
        ("Trace: Source", 30),
        ("Trace: Method", 35),
        ("Trace: Intel Report Link", 25),
        ("Trace: Derivation", 45),
    ]
    
    header_row = 3
    for col_idx, (h, _) in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = _FONT_VC_HEADER
        c.fill = _FILL_VC_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
    
    # Traceability sub-headers (columns 16-19) get olive green
    for col_idx in range(17, 21):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = _FONT_TRACE_HEADER
        c.fill = _FILL_VC_TRACE_HEADER
    
    # == Data rows ==
    current_group = None
    write_row = 4
    for r in rows:
        row_idx = write_row
        group = r.get("group", "")
        field_num = r.get("field_number", "")
        status = r.get("status", "")
        req = r.get("requirement", "")
        
        # Group separator row when group changes
        if group != current_group:
            current_group = group
            # Insert a group header row
            group_fill = _VC_GROUP_HEADER_FILLS.get(group, _FILL_VC_HEADER)
            group_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=20)
            group_cell = ws.cell(row=row_idx, column=1, value=f"  >> {group}")
            group_cell.font = group_font
            group_cell.fill = group_fill
            group_cell.alignment = Alignment(horizontal="left", vertical="center")
            group_cell.border = _BORDER_THIN
            for c in range(2, 21):
                ws.cell(row=row_idx, column=c).fill = group_fill
                ws.cell(row=row_idx, column=c).border = _BORDER_THIN
            ws.row_dimensions[row_idx].height = 22
            write_row += 1
            row_idx = write_row
        
        # Row fill based on group
        row_fill = _VC_GROUP_ROW_FILLS.get(group)
        
        c = ws.cell(row=row_idx, column=1, value=field_num)
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Group
        c = ws.cell(row=row_idx, column=2, value=group)
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Field name (bold for Required)
        c = ws.cell(row=row_idx, column=3, value=r.get("field_name", ""))
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        if req in ("Req", "Req*"):
            c.font = Font(name="Calibri", size=10, bold=True)
        
        # Statement section
        c = ws.cell(row=row_idx, column=4, value=r.get("statement_section", ""))
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Requirement (color-coded)
        c = ws.cell(row=row_idx, column=5, value=req)
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        if req in ("Req", "Req*"):
            c.fill = _FILL_REQ
            c.font = Font(name="Calibri", size=9, bold=True, color="006100")
        elif req == "Opt":
            c.fill = _FILL_OPT
            c.font = Font(name="Calibri", size=9, color="4472C4")
        
        # Both years
        c = ws.cell(row=row_idx, column=6, value=r.get("both_years", ""))
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Type
        c = ws.cell(row=row_idx, column=7, value=r.get("field_type", ""))
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Engine use
        c = ws.cell(row=row_idx, column=8, value=r.get("engine_use", ""))
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        
        # Current value (number formatting)
        c = ws.cell(row=row_idx, column=9, value=r.get("current_value", ""))
        c.alignment = _ALIGN_RIGHT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        # Try to format as number
        cur_val = r.get("current_value", "")
        if cur_val and cur_val.replace(",", "").replace(".", "").replace("-", "").isdigit():
            try:
                c.value = float(cur_val.replace(",", ""))
                c.number_format = '#,##0.00;[Red](#,##0.00)'
            except (ValueError, TypeError):
                pass
        
        # Previous value
        c = ws.cell(row=row_idx, column=10, value=r.get("previous_value", ""))
        c.alignment = _ALIGN_RIGHT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
        prev_val = r.get("previous_value", "")
        if prev_val and prev_val.replace(",", "").replace(".", "").replace("-", "").isdigit():
            try:
                c.value = float(prev_val.replace(",", ""))
                c.number_format = '#,##0.00;[Red](#,##0.00)'
            except (ValueError, TypeError):
                pass
        
        # Status (color-coded)
        # YoY Growth %
        c = ws.cell(row=row_idx, column=11, value=r.get("yoy_growth", ""))
        c.alignment = _ALIGN_RIGHT
        c.border = _BORDER_THIN
        if row_fill:
            c.fill = row_fill
            
        # Status (color-coded)
        c = ws.cell(row=row_idx, column=12, value=status)
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        if status == "FOUND":
            c.fill = _FILL_FOUND
            c.font = _FONT_FOUND
        elif status == "NOT DISCLOSED":
            c.fill = _FILL_NOT_DISCLOSED
            c.font = _FONT_NOT_DISCLOSED
        elif status == "NOT APPLICABLE":
            c.fill = _FILL_NOT_APPLICABLE
            c.font = _FONT_NOT_APPLICABLE
        
        # Source statement
        c = ws.cell(row=row_idx, column=13, value=r.get("source_statement", ""))
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        
        # Source page
        c = ws.cell(row=row_idx, column=14, value=r.get("source_page", ""))
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        
        # Note no
        c = ws.cell(row=row_idx, column=15, value=r.get("note_no", ""))
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        
        # Raw line item
        c = ws.cell(row=row_idx, column=16, value=r.get("raw_line_item", ""))
        c.alignment = _ALIGN_LEFT
        c.border = _BORDER_THIN
        
        # == Traceability columns (16-19) ==
        for col_idx, key in enumerate(
            ["trace_source", "trace_method", "trace_intel_link", "trace_derivation"], 17
        ):
            c = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ""))
            c.alignment = _ALIGN_TOP_LEFT
            c.border = _BORDER_THIN
            c.fill = _FILL_TRACE
            c.font = Font(name="Calibri", size=9, color="595959")
    
    # == Column widths ==
    col_widths = [5, 14, 38, 28, 6, 8, 8, 30, 18, 18, 16, 22, 12, 8, 25, 30, 35, 25, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze panes (freeze header + title rows, and first 3 columns)
    ws.freeze_panes = "D4"
    
    # Auto-filter
    last_data_row = 3 + len(rows) + len(_VC_GROUP_HEADER_FILLS)  # account for group separator rows
    ws.auto_filter.ref = f"A3:S{last_data_row}"


# ===================================================================
# Category Sheet Builder
# ===================================================================

def _build_category_sheet(
    wb: "openpyxl.Workbook",
    category_name: str,
    sections: list[dict],
    text_extractions: list[dict],
    table_extractions: list[dict]
) -> None:
    # Ensure sheet name is valid
    invalid_chars = r"[]:*?/"
    safe_name = "".join(c for c in category_name if c not in invalid_chars)
    safe_name = safe_name[:31]
    ws = wb.create_sheet(title=safe_name)

    current_row = 1

    for section in sections:
        section_name = section.get("section_name", "Unknown Section")
        
        # Print Section Title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        cell = ws.cell(row=current_row, column=1, value=f"{section_name}")
        cell.font = _FONT_TITLE
        cell.fill = _FILL_SECTION
        cell.alignment = _ALIGN_LEFT
        current_row += 2

        if section.get("content_type") == "table":
            table_match = _find_table(section, table_extractions)
            if table_match:
                current_row = _render_table_block(ws, current_row, table_match)
            else:
                ws.cell(row=current_row, column=1, value="(Table data not extracted)").font = _FONT_DATA
                current_row += 2
        else:
            text_match = _find_text(section, text_extractions)
            if text_match:
                current_row = _render_text_block(ws, current_row, text_match)
            else:
                ws.cell(row=current_row, column=1, value="(Text data not extracted)").font = _FONT_DATA
                current_row += 2
                
        current_row += 2

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20


def _find_table(section: dict, table_extractions: list[dict]) -> dict | None:
    sec_name = section.get("section_name", "")
    
    # 1. Check exact match for generic table
    for table in table_extractions:
        if table.get("table_name") == sec_name:
            return table
            
    # 2. Check legacy map
    LEGACY_MAP = {
        "Standalone Balance Sheet": "standalone_balance_sheet",
        "Standalone Profit & Loss": "standalone_profit_and_loss",
        "Standalone Cash Flow": "standalone_cash_flow",
        "Consolidated Balance Sheet": "consolidated_balance_sheet",
        "Consolidated Profit & Loss": "consolidated_profit_and_loss",
        "Consolidated Cash Flow": "consolidated_cash_flow",
    }
    
    if sec_name in LEGACY_MAP:
        mapped = LEGACY_MAP[sec_name]
        for table in table_extractions:
            if table.get("table_name") == mapped:
                return table
                
    return None

def _find_text(section: dict, text_extractions: list[dict]) -> str | None:
    cat = section.get("category")
    subcat = section.get("section_name")
    for ext in text_extractions:
        if ext.get("category") == cat and ext.get("subcategory") == subcat:
            return ext.get("extracted_text")
    return None


def _render_text_block(ws: "openpyxl.worksheet.worksheet.Worksheet", start_row: int, text: str) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.font = _FONT_DATA
    cell.alignment = _ALIGN_TOP_LEFT
    
    # Estimate height (roughly 90 chars per line across 10 cols)
    lines = len(text.splitlines()) + (len(text) // 90)
    ws.row_dimensions[start_row].height = max(15, lines * 15)
    
    return start_row + 1

def _render_table_block(ws: "openpyxl.worksheet.worksheet.Worksheet", start_row: int, table_extract: dict) -> int:
    table_json = table_extract.get("table_json", {})
    if not table_json:
        ws.cell(row=start_row, column=1, value="No table data").font = _FONT_DATA
        return start_row + 1
        
    current_row = start_row

    title = table_json.get("title", "")
    if title:
        ws.cell(row=current_row, column=1, value=title).font = _FONT_SUBTITLE
        current_row += 1

    currency = table_json.get("currency", "")
    if currency:
        ws.cell(row=current_row, column=1, value=f"({currency})").font = _FONT_DATA
        current_row += 1

    # Legacy tables have dicts in rows, Generic tables have arrays
    rows = table_json.get("rows", [])
    if not rows:
        return current_row
        
    if isinstance(rows[0], list):
        # GENERIC TABLE
        headers = table_json.get("column_headers", [])
        if headers:
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = _FONT_HEADER
                cell.fill = _FILL_HEADER
                cell.alignment = _ALIGN_CENTER
                cell.border = _BORDER_THIN
            current_row += 1
            
        for row_data in rows:
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = _FONT_DATA
                cell.border = _BORDER_THIN
            current_row += 1
    else:
        # LEGACY FINANCIAL STATEMENT
        col_headers = table_json.get("column_headers", ["Note No.", "Current Period", "Previous Period"])
        has_note = any(h.lower().replace(".", "").replace(" ", "") == "noteno" for h in col_headers)
        headers = ["Particulars"] + col_headers

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
        current_row += 1

        prev_section_parts = []
        for row_data in rows:
            section = row_data.get("section")
            line_item = row_data.get("line_item", "")
            note_no = row_data.get("note_no")
            values = row_data.get("values", {})
            current_val = values.get("current_period")
            previous_val = values.get("previous_period")

            is_total = "total" in line_item.lower()

            section_parts = section.split(" > ") if section else []
            for depth, part in enumerate(section_parts):
                if depth < len(prev_section_parts) and prev_section_parts[depth] == part:
                    continue
                for insert_depth in range(depth, len(section_parts)):
                    p = section_parts[insert_depth]
                    if not p:
                        continue
                    indent = "    " * insert_depth
                    cell = ws.cell(row=current_row, column=1, value=f"{indent}{p}")
                    cell.font = _FONT_SECTION if insert_depth == 0 else _FONT_SUBSECTION
                    cell.border = _BORDER_THIN
                    for c in range(2, len(headers) + 1):
                        ws.cell(row=current_row, column=c).border = _BORDER_THIN
                    current_row += 1
                break
            prev_section_parts = section_parts

            indent = "    " * len(section_parts)
            display_label = f"{indent}{line_item}" if line_item else ""

            cell = ws.cell(row=current_row, column=1, value=display_label)
            cell.font = _FONT_TOTAL if is_total else _FONT_DATA
            cell.border = _BORDER_THIN
            if is_total:
                cell.fill = _FILL_TOTAL

            col_offset = 2
            if has_note:
                cell_note = ws.cell(row=current_row, column=col_offset, value=note_no)
                cell_note.alignment = _ALIGN_CENTER
                cell_note.border = _BORDER_THIN
                col_offset += 1

            for val in (current_val, previous_val):
                cell_val = ws.cell(row=current_row, column=col_offset)
                if val is not None:
                    cell_val.value = val
                    cell_val.number_format = '#,##0.00;[Red](#,##0.00)'
                else:
                    cell_val.value = "-"
                    cell_val.alignment = _ALIGN_CENTER
                cell_val.font = _FONT_TOTAL if is_total else _FONT_DATA
                cell_val.border = _BORDER_THIN
                if is_total:
                    cell_val.fill = _FILL_TOTAL
                col_offset += 1

            current_row += 1

    return current_row


# ===================================================================
# Existing Base Sheets (Metadata, Master Sections, Inventory)
# ===================================================================

def _build_metadata_sheet(wb: "openpyxl.Workbook", extraction_result: dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Metadata")
    _FONT_META_KEY = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    _FONT_META_VAL = Font(name="Calibri", size=10)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    cell = ws.cell(row=1, column=1, value="Extraction Metadata")
    cell.font = _FONT_TITLE
    
    row = 3
    metadata = extraction_result.get("metadata", {})
    for key, label in [
        ("file_name", "File Name"),
        ("financial_year", "Financial Year"),
        ("page_count", "Page Count"),
        ("extraction_timestamp", "Timestamp")
    ]:
        ws.cell(row=row, column=1, value=label).font = _FONT_META_KEY
        ws.cell(row=row, column=2, value=str(metadata.get(key, ""))).font = _FONT_META_VAL
        row += 1
        
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

def _build_master_sections_sheet(wb: "openpyxl.Workbook", master_sections: list[dict]) -> None:
    ws = wb.create_sheet(title="Master Sections")
    headers = ["Section ID", "Section Name", "Category", "Subcategory", "Start Page", "End Page",
               "Page Count", "Content Type", "Strategy", "Source", "Confidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
    
    for row_idx, section in enumerate(master_sections, start=2):
        ws.cell(row=row_idx, column=1, value=section.get("section_id"))
        ws.cell(row=row_idx, column=2, value=section.get("section_name"))
        ws.cell(row=row_idx, column=3, value=section.get("category"))
        ws.cell(row=row_idx, column=4, value=section.get("subcategory", ""))
        ws.cell(row=row_idx, column=5, value=section.get("start_page"))
        ws.cell(row=row_idx, column=6, value=section.get("end_page"))
        ws.cell(row=row_idx, column=7, value=section.get("page_count",
                    section.get("end_page", 0) - section.get("start_page", 0) + 1))
        ws.cell(row=row_idx, column=8, value=section.get("content_type"))
        ws.cell(row=row_idx, column=9, value=section.get("extraction_strategy"))
        ws.cell(row=row_idx, column=10, value=section.get("source", "taxonomy"))
        ws.cell(row=row_idx, column=11, value=section.get("confidence"))
        
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.freeze_panes = "A2"

def _build_table_inventory_sheet(wb: "openpyxl.Workbook", inventory: list[dict]) -> None:
    ws = wb.create_sheet(title="Table Inventory")
    headers = ["Table ID", "Table Name", "Table Category", "Page No.", "Needs VLM", "Parent Section"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        
    for row_idx, item in enumerate(inventory, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("table_id"))
        ws.cell(row=row_idx, column=2, value=item.get("table_name"))
        ws.cell(row=row_idx, column=3, value=item.get("table_category", "other"))
        ws.cell(row=row_idx, column=4, value=item.get("page_no"))
        ws.cell(row=row_idx, column=5, value=str(item.get("needs_vlm", False)))
        ws.cell(row=row_idx, column=6, value=item.get("parent_section_id", ""))
        
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["F"].width = 25
    ws.freeze_panes = "A2"
