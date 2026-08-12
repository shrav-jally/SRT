"""Excel Exporter for CustomExtractionResultDocument.

Exports custom spec extraction results to presentation-friendly OpenPyXL workbooks.
"""

from __future__ import annotations

import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from contracts import CustomExtractionResultDocument, ExtractionStatus


def export_custom_extraction_to_excel(
    result_doc: CustomExtractionResultDocument,
    output_path: Path | str | None = None,
) -> bytes:
    """Export CustomExtractionResultDocument to Excel file or bytes.

    Parameters
    ----------
    result_doc : CustomExtractionResultDocument
        The custom spec extraction output document.
    output_path : Path | str, optional
        Target file path. If None, returns raw Excel bytes.

    Returns
    -------
    bytes
        Excel file bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom Extraction Spec"

    # Styling Token Tokens
    font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

    fill_found = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # Soft Green
    fill_not_found = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Soft Red/Orange
    fill_ambiguous = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Soft Yellow

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title block
    ws.cell(row=1, column=1, value=f"Custom Spec Extraction Report — Document ID: {result_doc.document_id}").font = font_title
    ws.row_dimensions[1].height = 25

    headers = [
        "Category",
        "Sub Category",
        "Entity Name",
        "Entity Type",
        "Extraction Mode",
        "Status",
        "Raw Value",
        "Normalized Value",
        "Unit",
        "Currency",
        "Confidence",
        "Explanation",
        "Page Number(s)",
        "Section ID",
        "Table ID",
        "Cell ID",
    ]

    # Header Row
    ws.row_dimensions[3].height = 24
    for c_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data Rows
    row_num = 4
    for res in result_doc.results:
        ws.row_dimensions[row_num].height = 20

        # Provenance summary
        pages_str = ", ".join(str(p.page_number) for p in res.provenance) if res.provenance else "-"
        sec_id_str = ", ".join(str(p.section_id) for p in res.provenance if p.section_id) or "-"
        tbl_id_str = ", ".join(str(p.table_id) for p in res.provenance if p.table_id) or "-"
        cell_id_str = ", ".join(str(p.cell_id) for p in res.provenance if p.cell_id) or "-"

        row_vals = [
            res.category,
            res.subcategory,
            res.entity_name,
            res.entity_type,
            res.extraction_mode.value,
            res.status.value,
            res.value_raw or "-",
            str(res.value_normalized) if res.value_normalized is not None else "-",
            res.unit or "-",
            res.currency or "-",
            f"{res.confidence * 100:.0f}%",
            res.explanation or "-",
            pages_str,
            sec_id_str,
            tbl_id_str,
            cell_id_str,
        ]

        # Determine row status fill
        if res.status == ExtractionStatus.FOUND:
            status_fill = fill_found
        elif res.status in (ExtractionStatus.NOT_FOUND, ExtractionStatus.FAILED_VALIDATION):
            status_fill = fill_not_found
        else:
            status_fill = fill_ambiguous

        for c_idx, val in enumerate(row_vals, start=1):
            c_cell = ws.cell(row=row_num, column=c_idx, value=val)
            c_cell.border = border_thin
            c_cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Highlight status cell
            if c_idx == 6:  # Status column
                c_cell.fill = status_fill
                c_cell.font = Font(bold=True)
                c_cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    bio = io.BytesIO()
    wb.save(bio)
    excel_bytes = bio.getvalue()

    if output_path:
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "wb") as f:
            f.write(excel_bytes)

    return excel_bytes
